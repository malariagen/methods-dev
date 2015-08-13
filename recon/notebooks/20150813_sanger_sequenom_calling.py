# -*- coding: utf-8 -*-
# <nbformat>3.0</nbformat>

# <headingcell level=1>

# Setup

# <codecell>

%run _shared_setup.ipynb

import openpyxl
import locale

# <codecell>

# sanger_sequenom_fn = '/nfs/team112_internal/rp7/recon/sanger_sequenom/processed_data/DK_KR_iPLEX_DK1079_W1381.txt'
plate_name = 'W36000_QC409376_409377_409378_409379_20150715'
data_dir = '/nfs/team112_internal/rp7/recon/sanger_sequenom'
sanger_sequenom_format = '%s/combined_data/%s.txt'
calling_parameters_fn = '/nfs/team112_internal/rp7/recon/sanger_sequenom/processed_data/DK1079_W1381_sanger_test/data_DK1079_W1381_parameters.txt'
sample_mappings_fn = '/nfs/team112_internal/rp7/recon/sanger_sequenom/sample_mappings/DK1079_sanger_ids_oxford_codes.xlsx'
kirk_calls_fn = ''
illumina_vcf_fn = PGV4_VCF_FN
ref_genome_fn = REF_GENOME
water_names=['WATER', 'X', '']
cache_dir_format = '%s/cache_data/%s'
calls_dir_format = '%s/genotype_calls/%s'
plot_dir_format = '%s/plots/%s'
summary_file_format = '%s/summary/%s'

# <codecell>

genotype_colours = ['orange', 'lightgray', 'darkred', 'darkgreen', 'darkblue']
# genotype_colours_water = ['black', 'lightgray', 'red', 'green', 'blue']
genotype_colours_water = ['black', 'lightgray', 'fuchsia', 'chartreuse', 'cyan']
named_genotype_colours = collections.OrderedDict()
named_genotype_colours['./.'] = 'orange'
named_genotype_colours['?'] = 'lightgray'
named_genotype_colours['0/0'] = 'darkred'
named_genotype_colours['1/1'] = 'darkgreen'
named_genotype_colours['0/1'] = 'darkblue'
named_genotype_colours_water = collections.OrderedDict()
named_genotype_colours_water['./.'] = 'black'
named_genotype_colours_water['?'] = 'lightgray'
# named_genotype_colours_water['0/0'] = 'red'
# named_genotype_colours_water['1/1'] = 'green'
# named_genotype_colours_water['0/1'] = 'blue'
named_genotype_colours_water['0/0'] = 'fuchsia'
named_genotype_colours_water['1/1'] = 'chartreuse'
named_genotype_colours_water['0/1'] = 'cyan'


# <codecell>

traffic_light_colours = collections.OrderedDict()
traffic_light_colours['10'] = 'chartreuse'
traffic_light_colours['8-9'] = 'green'
traffic_light_colours['4-7'] = 'orange'
traffic_light_colours['1-3'] = 'darkred'
traffic_light_colours['0'] = 'red'

# <codecell>


# <headingcell level=1>

# Functions

# <codecell>

def determine_GT(AD):
    if sum(AD) < 5:
        return("./.")
    if len(AD) == 3:
        if AD[0] > 1 and AD[1] > 1:
            return("0/1")
        if AD[0] > 1 and AD[2] > 1:
            return("0/2")
        if AD[1] > 1 and AD[2] > 1:
            return("1/2")
    if(AD[0] > 1 and AD[1] > 1):
        return("0/1")
    if(AD[0] <= 1 and AD[1] > 1):
        return("1/1")
    if(AD[0] > 1 and AD[1] <= 1):
        return("0/0")
    else:
        return("unknown")
    

# <codecell>

def calc_low_intensity(x=3, y=2, theta=45):
    if x is None or x==0:
        return None
    else:
        tan_sq_theta = math.tan(math.radians(theta))**2
        return(math.sqrt((tan_sq_theta + 1) * (((x**2)*(y**2))/(((x**2)*tan_sq_theta) + y**2))))

# <codecell>

def determine_genotype(rec):
    if rec['INTENSITY'] is None or rec['THETA'] is None or rec['low_intensity_cutoff'] is None or rec['genotype_threshold_degrees_1'] is None or rec['genotype_threshold_degrees_2'] is None:
        return(None)
    elif rec['INTENSITY'] < rec['low_intensity_cutoff']:
        return('./.')
    elif rec['THETA'] < rec['genotype_threshold_degrees_1']:
        return('0/0')
    elif rec['THETA'] > rec['genotype_threshold_degrees_2']:
        return('1/1')
    else:
        return('0/1')

# <codecell>

def determine_species(rec):
    assay_split=rec['ASSAY_ID'].split('_')
    species_identifier = assay_split[0].split('-')[1]
    if(species_identifier=='Pf'):
        return('Pf')
    elif(species_identifier=='gb'):
        return('Pv')
    elif(species_identifier=='amelogenin'):
        return('Hs')
    
def determine_chrom(rec):
    assay_split=rec['ASSAY_ID'].split('_')
    species_identifier = assay_split[0].split('-')[1]
    if(species_identifier=='Pf'):
        return('_'.join((assay_split[3], assay_split[4], assay_split[5])))
    elif(species_identifier=='gb'):
        return('_'.join((species_identifier, assay_split[1])))
    elif(species_identifier=='amelogenin'):
        return('XY')
    
def determine_pos(rec):
    assay_split=rec['ASSAY_ID'].split('_')
    species_identifier = assay_split[0].split('-')[1]
    if(species_identifier=='Pf'):
        return(int(assay_split[6]))
    elif(species_identifier=='gb'):
        return(int(assay_split[2]))
    elif(species_identifier=='amelogenin'):
        return(0)
    
def determine_ref(rec, ref_dict):
    if(rec['species']=='Pf'):
        return(ref_dict[rec['chrom']].seq[rec['pos']-1])
    else:
        return(None)
    
def determine_alt(rec):
    if(rec['ref'] is None):
        return(None)
    elif(rec['ref'] == rec['ALLELE_1']):
        return(rec['ALLELE_2'])
    else:
        return(rec['ALLELE_1'])
    

# <codecell>

def cache_sample_mappings(sample_mappings_fn, plate_name, sanger_id_name='SANGER SAMPLE ID', recon_id_name='sample_code',
                          rewrite=False):
    cache_fn = "%s/tbl_sample_mappings" % (cache_dir_format % (data_dir, plate_name))
    if not os.path.exists(os.path.dirname(cache_fn)):
        os.mkdir(os.path.dirname(cache_fn))
    if not os.path.exists(cache_fn) or rewrite:
        if sample_mappings_fn.endswith('.xlsx'):
            tbl_sample_mappings = (etl
                .fromxlsx(sample_mappings_fn, data_only=True)
                .rename(sanger_id_name, 'sanger_id')
                .rename(recon_id_name, 'recon_id')
                .cut(['sanger_id', 'recon_id'])
            )
            tbl_sample_mappings.topickle(cache_fn)
        elif sample_mappings_fn.endswith('.xls'):
            tbl_sample_mappings = (etl
                .fromxls(sample_mappings_fn)
                .skip(8)
                .rename(sanger_id_name, 'sanger_id')
                .rename(recon_id_name, 'recon_id')
                .cut(['sanger_id', 'recon_id'])
            )
            tbl_sample_mappings.topickle(cache_fn)
    return etl.frompickle(cache_fn)
    

# <codecell>

def cache_raw_sequenom_data(plate_name, tbl_sample_mappings, ref_genome, rewrite=False):
    ref_dict = SeqIO.to_dict(SeqIO.parse(open(ref_genome), "fasta"))
    sanger_sequenom_fn = sanger_sequenom_format % (data_dir, plate_name)
    cache_fn = "%s/tbl_sanger_sequenom" % (cache_dir_format % (data_dir, plate_name))
    if not os.path.exists(os.path.dirname(cache_fn)):
        os.mkdir(os.path.dirname(cache_fn))
    if not os.path.exists(cache_fn) or rewrite:
        tbl_sanger_sequenom = (etl
            .fromtsv(sanger_sequenom_fn)
            .convertnumbers()
            .leftjoin(tbl_sample_mappings, lkey='SAMPLE_ID', rkey='sanger_id')
            .addfield('species', determine_species)
            .addfield('chrom', determine_chrom)
            .addfield('pos', determine_pos)
            .addfield('ref', lambda rec: determine_ref(rec, ref_dict))
#             .addfield('alt', determine_alt)
            .sort(['PLATE', 'WELL_POSITION', 'ASSAY_ID', 'ALLELE'])
#             .sort(['PLATE', 'WELL_POSITION', 'ASSAY_ID', 'MASS'])
        )
        tbl_sanger_sequenom.topickle(cache_fn)
    return etl.frompickle(cache_fn)
    

# <codecell>

def cache_illumina(illumina_vcf_fn, tbl_sanger_sequenom, plate_name, rewrite=False):
    cache_fn = "%s/tbl_illumina" % (cache_dir_format % (data_dir, plate_name))
    if not os.path.exists(os.path.dirname(cache_fn)):
        os.mkdir(os.path.dirname(cache_fn))
    if not os.path.exists(cache_fn) or rewrite:
        tbl_illumina_samples = tbl_sanger_sequenom.distinct('recon_id').cut(['recon_id', 'species'])
        for i, (chrom, pos) in enumerate(tbl_sanger_sequenom.selecteq('species', 'Pf').distinct(('chrom', 'pos')).values(('chrom', 'pos'))):
            if i==0:
                tbl_illumina = (etl.fromvcf(illumina_vcf_fn, chrom=chrom, start=pos, stop=pos)
                    .vcfmeltsamples()
                    .vcfunpackcall()
                    .cutout('GT')
                    .addfield('GT', lambda rec: determine_GT(rec['AD']))
                    .join(tbl_illumina_samples, lkey='SAMPLE', rkey='recon_id')
                    .cut(['CHROM', 'POS', 'SAMPLE', 'GT', 'AD'])
                    .rename('GT', 'illumina_gt')
                )
            else:
                tbl_illumina = tbl_illumina.cat(etl.fromvcf(illumina_vcf_fn, chrom=chrom, start=pos, stop=pos)
                    .vcfmeltsamples()
                    .vcfunpackcall()
                    .cutout('GT')
                    .addfield('GT', lambda rec: determine_GT(rec['AD']))
                    .join(tbl_illumina_samples, lkey='SAMPLE', rkey='recon_id')
                    .cut(['CHROM', 'POS', 'SAMPLE', 'GT', 'AD'])
                    .rename('GT', 'illumina_gt')
                )
        tbl_illumina.topickle(cache_fn)
    return etl.frompickle(cache_fn)
                

# <codecell>

def cache_calling_parameters(calling_parameters_fn, plate_name, rewrite=False):
    cache_fn = "%s/tbl_calling_parameters" % (cache_dir_format % (data_dir, plate_name))
    if not os.path.exists(os.path.dirname(cache_fn)):
        os.mkdir(os.path.dirname(cache_fn))
    if not os.path.exists(cache_fn) or rewrite:
        tbl_calling_parameters_sep = etl.fromtsv(calling_parameters_fn).selectne('assay_code', 'amelogenin_XY_SNP1_E').convertnumbers()
        tbl_calling_parameters_1 = (tbl_calling_parameters_sep
            .addrownumbers()
            .select(lambda rec: rec['row'] % 2 == 1)
            .rename('allele', 'allele_1')
            .rename('lift_degrees', 'lift_degrees_1')
            .rename('genotype_threshold_degrees', 'genotype_threshold_degrees_1')
            .rename('low_intensity_cutoff', 'low_intensity_cutoff_1')
            .cut(['plate_code', 'assay_code', 'alleles_pair', 'allele_1', 'lift_degrees_1', 'genotype_threshold_degrees_1', 'low_intensity_cutoff_1'])
        )
        tbl_calling_parameters_2 = (tbl_calling_parameters_sep
            .addrownumbers()
            .select(lambda rec: rec['row'] % 2 == 0)
            .rename('allele', 'allele_2')
            .rename('lift_degrees', 'lift_degrees_2')
            .rename('genotype_threshold_degrees', 'genotype_threshold_degrees_2')
            .rename('low_intensity_cutoff', 'low_intensity_cutoff_2')
            .cut(['allele_2', 'lift_degrees_2', 'genotype_threshold_degrees_2', 'low_intensity_cutoff_2'])
        )
        tbl_calling_parameters = (
            tbl_calling_parameters_1
            .annex(tbl_calling_parameters_2)
            .convert('assay_code', lambda val: 'W36000-' + val)
        )
        tbl_calling_parameters.topickle(cache_fn)
    return etl.frompickle(cache_fn)
    

# <codecell>

def determine_sequenom_gt(rec):
    if(rec['ref'] is None):
        return('?')
    elif(rec['ref'] is None or rec['SEQUENOM_GENOTYPE'] is None or rec['SEQUENOM_GENOTYPE'] == ''):
        return('./.')
    elif(rec['SEQUENOM_GENOTYPE'] == rec['ref']):
        return('0/0')
    elif(rec['SEQUENOM_GENOTYPE'] == rec['alt']):
        return('1/1')
    elif(rec['SEQUENOM_GENOTYPE'] == rec['ref']+rec['alt'] or rec['SEQUENOM_GENOTYPE'] == rec['alt']+rec['ref']):
        return('0/1')
    else:
        return('error')
    

# <codecell>

def determine_called_genotype(rec):
    if rec['INTENSITY'] is None or rec['THETA'] is None or rec['low_intensity_cutoff'] is None or rec['genotype_threshold_degrees_1'] is None or rec['genotype_threshold_degrees_2'] is None:
        return('?')
    elif rec['INTENSITY'] < rec['low_intensity_cutoff']:
        return('')
    elif rec['THETA'] < rec['genotype_threshold_degrees_1']:
        return(rec['ALLELE_1'])
    elif rec['THETA'] > rec['genotype_threshold_degrees_2']:
        return(rec['ALLELE_2'])
    else:
        return(rec['ALLELE_1']+rec['ALLELE_2'])

# <codecell>

def determine_called_gt(rec):
    if rec['INTENSITY'] is None or rec['THETA'] is None or rec['low_intensity_cutoff'] is None or rec['genotype_threshold_degrees_1'] is None or rec['genotype_threshold_degrees_2'] is None:
        return('?')
    elif rec['INTENSITY'] < rec['low_intensity_cutoff']:
        return('./.')
    elif rec['THETA'] < rec['genotype_threshold_degrees_1']:
        if rec['ref'] is None or rec['ALLELE_1'] == rec['ref']:
            return('0/0')
        else:
            return('1/1')
    elif rec['THETA'] > rec['genotype_threshold_degrees_2']:
        if rec['ref'] is None or rec['ALLELE_1'] == rec['ref']:
            return('1/1')
        else:
            return('0/0')
    else:
        return('0/1')

# <codecell>

def determine_illumina_genotype(rec):
    if rec['illumina_gt'] == '0/0':
        return(rec['ref'])
    elif rec['illumina_gt'] == '1/1':
        return(rec['alt'])
    elif rec['illumina_gt'] == '0/1':
        return(rec['ref']+rec['alt'])
    elif rec['illumina_gt'] == './.':
        return('')
    else:
        return('?')

# <codecell>

def determine_theta(rec):
    if rec['HEIGHT_1'] == 0 and rec['HEIGHT_2'] == 0:
        return(np.float64(0.0))
    else:
        return(math.degrees(math.atan(np.float64(rec['HEIGHT_2'])/np.float64(rec['HEIGHT_1']))))

# <codecell>

def determine_intensity(rec):
    return(math.sqrt(math.pow(np.float64(rec['HEIGHT_1']), 2) + math.pow(np.float64(rec['HEIGHT_2']), 2)))

# <codecell>

def cache_calling_results(tbl_sanger_sequenom, tbl_calling_parameters, tbl_illumina, plate_name, ref_genome, rewrite=False):
    ref_dict = SeqIO.to_dict(SeqIO.parse(open(ref_genome), "fasta"))
    cache_fn = "%s/tbl_calling_results" % (cache_dir_format % (data_dir, plate_name))
    if not os.path.exists(os.path.dirname(cache_fn)):
        os.mkdir(os.path.dirname(cache_fn))
    if not os.path.exists(cache_fn) or rewrite:
        tbl_first_allele = (tbl_sanger_sequenom
            .addrownumbers()
            .select(lambda rec: rec['row'] % 2 == 1)
            .rename('ALLELE', 'ALLELE_1')
            .rename('HEIGHT', 'HEIGHT_1')
            .rename('MASS', 'MASS_1')
            .rename('GENOTYPE_ID', 'SEQUENOM_GENOTYPE')
            .cut(['ASSAY_ID', 'chrom', 'pos', 'ref', 'SEQUENOM_GENOTYPE', 'SAMPLE_ID', 'recon_id', 'species', 'STATUS', 'WELL_POSITION', 'ALLELE_1', 'HEIGHT_1', 'MASS_1'])
        )
        tbl_second_allele = (tbl_sanger_sequenom
            .addrownumbers()
            .select(lambda rec: rec['row'] % 2 == 0)
            .rename('ALLELE', 'ALLELE_2')
            .rename('HEIGHT', 'HEIGHT_2')
            .rename('MASS', 'MASS_2')
            .rename('GENOTYPE_ID', 'SEQUENOM_GENOTYPE')
            .cut(['ALLELE_2', 'HEIGHT_2', 'MASS_2'])
        )
        tbl_calling_results = (
            tbl_first_allele
            .annex(tbl_second_allele)
            .addfield('THETA', determine_theta)
            .addfield('INTENSITY', determine_intensity)
            .leftjoin(tbl_calling_parameters, lkey='ASSAY_ID', rkey='assay_code')
            .addfield('low_intensity_cutoff', lambda rec: calc_low_intensity(rec['low_intensity_cutoff_1'], rec['low_intensity_cutoff_2'], rec['THETA']))
            .addfield('alt', determine_alt)
            .addfield('sequenom_gt', determine_sequenom_gt)
            .addfield('called_genotype', determine_called_genotype)
            .addfield('called_gt', determine_called_gt)
            .leftjoin(tbl_illumina, lkey=('chrom', 'pos', 'recon_id'), rkey=('CHROM', 'POS', 'SAMPLE'))
            .replace('illumina_gt', None, '?')
            .addfield('illumina_genotype', determine_illumina_genotype)
        )
        tbl_calling_results.topickle(cache_fn)
    return etl.frompickle(cache_fn)
    

# <codecell>

def sequenom_scatter_plot(tbl_calling_results, tbl_calling_parameters, genotype_column='SEQUENOM_GENOTYPE', water_names=['WATER', 'X', ''],
                          plot_type='heights', plot_scale=0.0):
    if plot_type=='theta':
        x_col='THETA'
        y_col='INTENSITY'
    else:
        x_col='HEIGHT_1'
        y_col='HEIGHT_2'
    
    fig = figure(figsize=(12, 7.5))
    for i, ASSAY_ID in enumerate(tbl_calling_results
        .distinct('ASSAY_ID')
        .values('ASSAY_ID')
        .array()
    ):
        fields_of_interest = ['genotype_threshold_degrees_1', 'genotype_threshold_degrees_2',
                              'low_intensity_cutoff_1', 'low_intensity_cutoff_2']
        if len(tbl_calling_parameters.selecteq('assay_code', ASSAY_ID).data()) > 0:
            (gtd1, gtd2, lic1, lic2) = tbl_calling_parameters.selecteq('assay_code', ASSAY_ID).cut(fields_of_interest).data()[0]
        else:
            (gtd1, gtd2, lic1, lic2) = (0.0, 90.0, 0.0, 0.0)

        genotypes = (tbl_calling_results
            .selecteq('ASSAY_ID', ASSAY_ID)
            .distinct(genotype_column)
            .addfield('genotype_length', lambda rec: len(rec[genotype_column]))
            .sort('genotype_length')
            .values(genotype_column)
            .array()
        )
        if not ('' in genotypes):
            genotypes = np.insert(genotypes, 0, '')
        if not ('?' in genotypes):
            genotypes = np.insert(genotypes, 1, '?')

#         print(ASSAY_ID, genotypes)
        ax = fig.add_subplot(3, 5, i+1)
        max_x = 0
        max_y = 0
#         for j, genotype in enumerate(insert(genotypes, 0, None)):
        for j, genotype in enumerate(genotypes):
            tbl_calling_results_this_genotype = (tbl_calling_results
                .selecteq('ASSAY_ID', ASSAY_ID)
                .selecteq(genotype_column, genotype)
            )
#             print(tbl_calling_results_this_genotype.header())

# First plot waters
            x_heights = (tbl_calling_results_this_genotype
                .selectin('recon_id', water_names)
                .values(x_col)
                .array()
            )
            y_heights = (tbl_calling_results_this_genotype
                .selectin('recon_id', water_names)
                .values(y_col)
                .array()
            )
            if genotype in named_genotype_colours_water:
                genotype_colour=named_genotype_colours_water[genotype]
            else:
                genotype_colour=genotype_colours_water[j]
            ax.scatter(x_heights, y_heights, color=genotype_colour, alpha=0.5, label="%s (water)" % genotype)

            if len(x_heights) > 0 and np.nanmax(x_heights) > max_x:
                max_x = np.nanmax(x_heights)
            if len(y_heights) > 0 and np.nanmax(y_heights) > max_y:
                max_y = np.nanmax(y_heights)

# Then plot true samples
            x_heights = (tbl_calling_results_this_genotype
                .selectnotin('recon_id', water_names)
                .values(x_col)
                .array()
            )
            y_heights = (tbl_calling_results_this_genotype
                .selectnotin('recon_id', water_names)
                .values(y_col)
                .array()
            )
            if genotype in named_genotype_colours:
                genotype_colour=named_genotype_colours[genotype]
            else:
                genotype_colour=genotype_colours[j]
            ax.scatter(x_heights, y_heights, color=genotype_colour, alpha=0.5, label=genotype)

            if len(x_heights) > 0 and np.nanmax(x_heights) > max_x:
                max_x = np.nanmax(x_heights)
            if len(y_heights) > 0 and np.nanmax(y_heights) > max_y:
                max_y = np.nanmax(y_heights)

        first_row = (tbl_calling_results
            .selecteq('ASSAY_ID', ASSAY_ID)
            .head(1)
        )
        chrom = first_row.values('chrom')[0]
        pos = first_row.values('pos')[0]
        species = first_row.values('species')[0]
        allele_1 = first_row.values('ALLELE_1')[0] 
        allele_2 = first_row.values('ALLELE_2')[0] 
        ref = first_row.values('ref')[0]
        ref_alt_1 = 'ref' if allele_1==ref else 'alt'
        ref_alt_2 = 'ref' if allele_2==ref else 'alt'
        if species == 'Pf':
            x_label = "%s (%s)" % (allele_1, ref_alt_1)
            y_label = "%s (%s)" % (allele_2, ref_alt_2)
        else:
            x_label = allele_1
            y_label = allele_2
        ax.set_xlabel(x_label)
        ax.set_ylabel(y_label)
        
        assay_title = "%s %d" % (chrom, pos)          
        ax.set_title(assay_title)
        
#         Draw threshold lines
        if plot_type=='theta':
            ax.axvline(gtd1, color='red')
            ax.axvline(gtd2, color='green')

            li_x = np.linspace(0, 90, 100)
            li_y = np.array([calc_low_intensity(lic1, lic2, theta) for theta in li_x])

            ax.plot(li_x, li_y, color='black', linestyle='-', linewidth=1)
            
            if plot_scale > 0.0: # Use 0.0 to show full range of points
                ax.set_xlim(0, 90)
                ax.set_ylim(0, plot_scale*max(lic1, lic2))
            else:
                ax.set_xlim(0, 90)
                ax.set_ylim(bottom=0)

        else:
            max_val = max(max_x, max_y)
            ax.plot([0, max_val], [0, max_val*math.tan(math.radians(gtd1))], color='red', linestyle='-', linewidth=1)
            ax.plot([0, max_val*math.tan(math.radians(90.0-gtd2))], [0, max_val], color='green', linestyle='-', linewidth=1)

            li_x = np.linspace(0, lic1, 100)
            li_y = np.sqrt( (lic2**2) * (1 - ( (li_x**2) / (lic1**2) ) ) )
            ax.plot(li_x, li_y, color='black', linestyle='-', linewidth=1)

            if plot_scale > 0.0: # Use 0.0 to show full range of points
                ax.set_xlim(0, plot_scale*lic1)
                ax.set_ylim(0, plot_scale*lic2)
            else:
                ax.set_xlim(0.0, max_val*1.1)
                ax.set_ylim(0.0, max_val*1.1)
        
#         ax.legend(ncol=2)

    fig.tight_layout()

# <codecell>

def determine_traffic_light(rec):
    if rec['# calls'] == 10:
        return('10')
    elif rec['# calls'] in [8, 9]:
        return('8-9')
    elif rec['# calls'] in [4, 5, 6, 7]:
        return('4-7')
    elif rec['# calls'] in [1, 2, 3]:
        return('1-3')
    elif rec['# calls'] == 0:
        return('0')
    

# <codecell>

def traffic_light_plot(tbl_calling_results, traffic_light_colours=traffic_light_colours):
    well_aggregation = collections.OrderedDict()
    well_aggregation['# calls'] = 'called_gt', lambda rec: np.sum(x in ['0/0', '0/1', '1/1'] for x in list(rec))
    well_aggregation['recon_id'] = 'recon_id', lambda rec: list(rec)[0]

    tbl_calls_per_well = (tbl_calling_results
        .selecteq('species', 'Pf')
        .aggregate('WELL_POSITION', well_aggregation)
        .addfield('row_number', lambda rec: ord(rec['WELL_POSITION'][0]) - 64)
        .addfield('column_number', lambda rec: int(rec['WELL_POSITION'][1:3]))
        .addfield('traffic_light', determine_traffic_light)
    )
    fig = figure(figsize=(12, 7.5))
    ax = fig.add_subplot(1, 1, 1)
    
#     Plot all wells
    for traffic_light_colour in traffic_light_colours:
        x_vals = tbl_calls_per_well.selecteq('traffic_light', traffic_light_colour).values('column_number').array()
        y_vals = 16 - tbl_calls_per_well.selecteq('traffic_light', traffic_light_colour).values('row_number').array()
        
        ax.scatter(x_vals, y_vals, color=traffic_light_colours[traffic_light_colour], s=100, label=traffic_light_colour)
    
#     Highlight waters
    x_vals = tbl_calls_per_well.selectin('recon_id', water_names).values('column_number').array()
    y_vals = 16 - tbl_calls_per_well.selectin('recon_id', water_names).values('row_number').array()

    ax.scatter(x_vals, y_vals, facecolors='none', edgecolors='blue', marker='o', s=400, label='WATER')

    ax.xaxis.set_ticks_position('top')
    ax.yaxis.set_ticks_position('left')

    ax.set_xlim(0.5, 24.5)
    ax.set_ylim(-0.5, 15.5)
    
    ax.set_xticks(range(1, 25))
    ax.set_yticks(range(0, 16))
    ax.set_yticklabels(['P', 'O', 'N', 'M', 'L', 'K', 'J', 'I', 'H', 'G', 'F', 'E', 'D', 'C', 'B', 'A'])
    
    ax.legend(loc='upper left', bbox_to_anchor=(1, 1), labelspacing=2, title='Number of Pf calls')

# Tight layout removes legend as outside axes so don't use here
    #     fig.tight_layout()

# <codecell>

def determine_results_summary(tbl_calling_results):
    results_summary = collections.OrderedDict()
    results_summary['a) # wells'] = len(tbl_calling_results.distinct('WELL_POSITION').data())
    results_summary['b) # samples'] = len(tbl_calling_results.distinct('WELL_POSITION').selectnotin('recon_id', water_names).data())
    results_summary['c) # sequenced samples'] = len(tbl_calling_results.distinct('WELL_POSITION').selectne('illumina_gt', '?').data())
    results_summary['d) # waters'] = len(tbl_calling_results.distinct('WELL_POSITION').selectin('recon_id', water_names).data())
    results_summary['e) # Pf calls samples'] = len(
        tbl_calling_results
        .selecteq('species', 'Pf')
        .selectnotin('recon_id', water_names)
        .selectin('called_gt', ['0/0', '0/1', '1/1'])
        .data()
    )
    results_summary['f) # Pf missing calls samples'] = len(
        tbl_calling_results
        .selecteq('species', 'Pf')
        .selectnotin('recon_id', water_names)
        .selectnotin('called_gt', ['0/0', '0/1', '1/1'])
        .data()
    )
    results_summary['g) # Pf calls water'] = len(
        tbl_calling_results
        .selecteq('species', 'Pf')
        .selectin('recon_id', water_names)
        .selectin('called_gt', ['0/0', '0/1', '1/1'])
        .data()
    )
    results_summary['h) # Pf missing calls water'] = len(
        tbl_calling_results
        .selecteq('species', 'Pf')
        .selectin('recon_id', water_names)
        .selectnotin('called_gt', ['0/0', '0/1', '1/1'])
        .data()
    )
    results_summary['i) % samples called'] = round(100.0*(results_summary['e) # Pf calls samples'] /
                                (results_summary['e) # Pf calls samples'] + results_summary['f) # Pf missing calls samples'])), 2)
    results_summary['j) % water missing'] = round(100.0*(results_summary['h) # Pf missing calls water'] /
                                (results_summary['h) # Pf missing calls water'] + results_summary['g) # Pf calls water'])), 2)

    return(etl
        .fromdicts([results_summary])
        .transpose()
        .pushheader(['variable', 'value'])
    )


def determine_calls_per_sample(tbl_calling_results):
    sample_aggregation = collections.OrderedDict()
    sample_aggregation['# calls'] = 'called_gt', lambda rec: np.sum(x in ['0/0', '0/1', '1/1'] for x in list(rec))

    tbl_calls_per_sample = (tbl_calling_results
        .selecteq('species', 'Pf')
        .selectnotin('recon_id', water_names)
        .aggregate('recon_id', sample_aggregation)
    )
    return(tbl_calls_per_sample)


def determine_snp_summary(tbl_calling_results):
    snp_aggregation = collections.OrderedDict()
    snp_aggregation['Chrom'] = 'chrom', lambda rec: list(rec)[0]
    snp_aggregation['Pos'] = 'pos', lambda rec: list(rec)[0]
    snp_aggregation['# Ref'] = ('recon_id', 'called_gt'), lambda rec: np.sum(not(x[0] in water_names) and x[1] == '0/0' for x in list(rec))
    snp_aggregation['# Het'] = ('recon_id', 'called_gt'), lambda rec: np.sum(not(x[0] in water_names) and x[1] == '0/1' for x in list(rec))
    snp_aggregation['# Alt'] = ('recon_id', 'called_gt'), lambda rec: np.sum(not(x[0] in water_names) and x[1] == '1/1' for x in list(rec))
    snp_aggregation['# No-call'] = ('recon_id', 'called_gt'), lambda rec: np.sum(not(x[0] in water_names) and x[1] == './.' for x in list(rec))
    snp_aggregation['# Water Ref'] = ('recon_id', 'called_gt'), lambda rec: np.sum((x[0] in water_names) and x[1] == '0/0' for x in list(rec))
    snp_aggregation['# Water Het'] = ('recon_id', 'called_gt'), lambda rec: np.sum((x[0] in water_names) and x[1] == '0/1' for x in list(rec))
    snp_aggregation['# Water Alt'] = ('recon_id', 'called_gt'), lambda rec: np.sum((x[0] in water_names) and x[1] == '1/1' for x in list(rec))
    snp_aggregation['# Water No-call'] = ('recon_id', 'called_gt'), lambda rec: np.sum((x[0] in water_names) and x[1] == './.' for x in list(rec))

    tbl_snp_summary = (tbl_calling_results
        .selecteq('species', 'Pf')
        .aggregate('ASSAY_ID', snp_aggregation)
    )

    return(tbl_snp_summary)


def determine_concordance_summary(tbl_calling_results):
    concordance_aggregation = collections.OrderedDict()
    concordance_aggregation['# Assays'] = len
    concordance_aggregation['# Concordant'] = ('called_gt', 'illumina_gt'), lambda rec: np.sum(x[0] != './.' and x[0] == x[1] for x in list(rec))
    concordance_aggregation['# Discordant'] = ('called_gt', 'illumina_gt'), lambda rec: np.sum(x[0] != './.' and x[1] != './.' and x[0] != x[1] for x in list(rec))
    concordance_aggregation['# Hom discordant'] = ('called_gt', 'illumina_gt'), lambda rec: np.sum((x[0] == '0/0' or x[0] == '1/1') and (x[1] == '0/0' or x[1] == '1/1') and x[0] != x[1] for x in list(rec))
    concordance_aggregation['# Het discordant'] = ('called_gt', 'illumina_gt'), lambda rec: np.sum((x[0] == '0/0' and x[1] == '0/1') or (x[0] == '1/1' and x[1] == '0/1') or (x[0] == '0/1' and x[1] == '0/0') or (x[0] == '0/1' and x[1] == '1/1') for x in list(rec))
    concordance_aggregation['# No call Sequenom'] = ('called_gt', 'illumina_gt'), lambda rec: np.sum(x[0] == './.' and x[1] != './.' for x in list(rec))
    concordance_aggregation['# No call Illumina'] = ('called_gt', 'illumina_gt'), lambda rec: np.sum(x[0] != './.' and x[1] == './.' for x in list(rec))
    concordance_aggregation['# No call both'] = ('called_gt', 'illumina_gt'), lambda rec: np.sum(x[0] == './.' and x[1] == './.' for x in list(rec))

    tbl_concordance_summary = (tbl_calling_results
        .selecteq('species', 'Pf')
        .selectne('illumina_gt', '?')
        .aggregate('plate_code', concordance_aggregation)
        .addfield('% Concordance', lambda rec: round((rec['# Concordant'] / (rec['# Concordant'] + rec['# Discordant'])) * 100, 2))
        .addfield('% Hom discordance', lambda rec: round((rec['# Hom discordant'] / (rec['# Concordant'] + rec['# Discordant'])) * 100, 2))
        .addfield('% Missing either assay', lambda rec: round(((rec['# No call Sequenom'] + rec['# No call Illumina'] + rec['# No call both']) / rec['# Assays']) * 100, 2))
    )

    return(tbl_concordance_summary)

def determine_snp_concordance_summary(tbl_calling_results):
    snp_concordance_aggregation = collections.OrderedDict()
    snp_concordance_aggregation['Chrom'] = 'chrom', lambda rec: list(rec)[0]
    snp_concordance_aggregation['Pos'] = 'pos', lambda rec: list(rec)[0]
    snp_concordance_aggregation['# Assays'] = len
    snp_concordance_aggregation['# Concordant'] = ('called_gt', 'illumina_gt'), lambda rec: np.sum(x[0] != './.' and x[0] == x[1] for x in list(rec))
    snp_concordance_aggregation['# Discordant'] = ('called_gt', 'illumina_gt'), lambda rec: np.sum(x[0] != './.' and x[1] != './.' and x[0] != x[1] for x in list(rec))
    snp_concordance_aggregation['# Hom discordant'] = ('called_gt', 'illumina_gt'), lambda rec: np.sum((x[0] == '0/0' or x[0] == '1/1') and (x[1] == '0/0' or x[1] == '1/1') and x[0] != x[1] for x in list(rec))
    snp_concordance_aggregation['# Het discordant'] = ('called_gt', 'illumina_gt'), lambda rec: np.sum((x[0] == '0/0' and x[1] == '0/1') or (x[0] == '1/1' and x[1] == '0/1') or (x[0] == '0/1' and x[1] == '0/0') or (x[0] == '0/1' and x[1] == '1/1') for x in list(rec))
    snp_concordance_aggregation['# No call Sequenom'] = ('called_gt', 'illumina_gt'), lambda rec: np.sum(x[0] == './.' and x[1] != './.' for x in list(rec))
    snp_concordance_aggregation['# No call Illumina'] = ('called_gt', 'illumina_gt'), lambda rec: np.sum(x[0] != './.' and x[1] == './.' for x in list(rec))
    snp_concordance_aggregation['# No call both'] = ('called_gt', 'illumina_gt'), lambda rec: np.sum(x[0] == './.' and x[1] == './.' for x in list(rec))

    tbl_snp_concordance_summary = (tbl_calling_results
        .selecteq('species', 'Pf')
        .selectne('illumina_gt', '?')
        .aggregate('ASSAY_ID', snp_concordance_aggregation)
        .addfield('% Concordance', lambda rec: round((rec['# Concordant'] / (rec['# Concordant'] + rec['# Discordant'])) * 100, 2))
        .addfield('% Hom discordance', lambda rec: round((rec['# Hom discordant'] / (rec['# Concordant'] + rec['# Discordant'])) * 100, 2))
        .addfield('% Missing either assay', lambda rec: round(((rec['# No call Sequenom'] + rec['# No call Illumina'] + rec['# No call both']) / rec['# Assays']) * 100, 2))
    )

    return(tbl_snp_concordance_summary.cutout('ASSAY_ID'))

def determine_sample_concordance_summary(tbl_calling_results):
    concordance_aggregation = collections.OrderedDict()
    concordance_aggregation['# Assays'] = len
    concordance_aggregation['# Concordant'] = ('called_gt', 'illumina_gt'), lambda rec: np.sum(x[0] != './.' and x[0] == x[1] for x in list(rec))
    concordance_aggregation['# Discordant'] = ('called_gt', 'illumina_gt'), lambda rec: np.sum(x[0] != './.' and x[1] != './.' and x[0] != x[1] for x in list(rec))
    concordance_aggregation['# Hom discordant'] = ('called_gt', 'illumina_gt'), lambda rec: np.sum((x[0] == '0/0' or x[0] == '1/1') and (x[1] == '0/0' or x[1] == '1/1') and x[0] != x[1] for x in list(rec))
    concordance_aggregation['# Het discordant'] = ('called_gt', 'illumina_gt'), lambda rec: np.sum((x[0] == '0/0' and x[1] == '0/1') or (x[0] == '1/1' and x[1] == '0/1') or (x[0] == '0/1' and x[1] == '0/0') or (x[0] == '0/1' and x[1] == '1/1') for x in list(rec))
    concordance_aggregation['# No call Sequenom'] = ('called_gt', 'illumina_gt'), lambda rec: np.sum(x[0] == './.' and x[1] != './.' for x in list(rec))
    concordance_aggregation['# No call Illumina'] = ('called_gt', 'illumina_gt'), lambda rec: np.sum(x[0] != './.' and x[1] == './.' for x in list(rec))
    concordance_aggregation['# No call both'] = ('called_gt', 'illumina_gt'), lambda rec: np.sum(x[0] == './.' and x[1] == './.' for x in list(rec))

    tbl_sample_concordance_summary = (tbl_calling_results
        .selecteq('species', 'Pf')
        .selectne('illumina_gt', '?')
        .aggregate('recon_id', concordance_aggregation)
        .addfield('% Concordance', lambda rec: round((rec['# Concordant'] / (rec['# Concordant'] + rec['# Discordant'])) * 100, 2))
        .addfield('% Hom discordance', lambda rec: round((rec['# Hom discordant'] / (rec['# Concordant'] + rec['# Discordant'])) * 100, 2))
        .addfield('% Missing either assay', lambda rec: round(((rec['# No call Sequenom'] + rec['# No call Illumina'] + rec['# No call both']) / rec['# Assays']) * 100, 2))
    )

    tbl_sample_concordance_summary = (tbl_sample_concordance_summary
        .selectne('# Concordant', 10)
        .sort('# No call Illumina', reverse=True)
        .sort('# No call Sequenom', reverse=True)
        .sort('% Concordance')
        .sort('% Hom discordance', reverse=True)
    )
    
    return(tbl_sample_concordance_summary)


def determine_concordance_matrix(tbl_calling_results, assay_id=None):
    if assay_id is None:
        tbl_concordance_matrix = (tbl_calling_results
            .selecteq('species', 'Pf')
            .selectne('illumina_gt', '?')
            .pivot('called_gt', 'illumina_gt', 'called_gt', len)
            .replaceall(None, 0)
        )
    else:
        tbl_concordance_matrix = (tbl_calling_results
            .selecteq('ASSAY_ID', assay_id)
            .selectne('illumina_gt', '?')
            .pivot('called_gt', 'illumina_gt', 'called_gt', len)
            .replaceall(None, 0)
        )
    return(tbl_concordance_matrix)


# <codecell>

def load_data(plate_name, calling_parameters_fn, sample_mappings_fn, ref_genome_fn, illumina_vcf_fn,
              recon_id_name='sample_code', rewrite=False):
    tbl_sample_mappings = cache_sample_mappings(sample_mappings_fn, plate_name, recon_id_name=recon_id_name, rewrite=rewrite)
    tbl_sanger_sequenom = cache_raw_sequenom_data(plate_name, tbl_sample_mappings, ref_genome_fn, rewrite=rewrite)
    tbl_illumina = cache_illumina(illumina_vcf_fn, tbl_sanger_sequenom, plate_name, rewrite=rewrite)
    tbl_calling_parameters = cache_calling_parameters(calling_parameters_fn, plate_name, rewrite=rewrite)
    tbl_calling_results = cache_calling_results(tbl_sanger_sequenom, tbl_calling_parameters, tbl_illumina,
                                                plate_name, ref_genome_fn, rewrite=rewrite)
    calls_fn = "%s/genotype_calls.txt" % (calls_dir_format % (data_dir, plate_name))
    if not os.path.exists(os.path.dirname(calls_fn)):
        os.makedirs(os.path.dirname(calls_fn))
    tbl_calling_results.totsv(calls_fn)
    return(tbl_calling_results)

# <codecell>

def create_summary_tables(plate_name, calling_parameters_fn, sample_mappings_fn, ref_genome_fn, illumina_vcf_fn):
    summary_fn = "%s/summary_%s.xlsx" % (summary_file_format % (data_dir, plate_name), plate_name)
    if not os.path.exists(os.path.dirname(summary_fn)):
        os.makedirs(os.path.dirname(summary_fn))
    tbl_calling_results = load_data(plate_name, calling_parameters_fn, sample_mappings_fn, ref_genome_fn, illumina_vcf_fn)

    wb = openpyxl.Workbook(optimized_write=True, encoding=locale.getpreferredencoding())

    tbl_results_summary = determine_results_summary(tbl_calling_results)
    ws = wb.create_sheet(title='summary')
    for row in tbl_results_summary:
        ws.append(row)
    
    tbl_calls_per_sample = determine_calls_per_sample(tbl_calling_results)
    ws = wb.create_sheet(title='calls per sample')
    for row in tbl_calls_per_sample.valuecounts('# calls').sort('# calls', reverse=True):
        ws.append(row)
    ws = wb.create_sheet(title='sample missingness')
    for row in tbl_calls_per_sample.selectlt('# calls', 10).sort('# calls', reverse=False):
        ws.append(row)

    tbl_snp_summary = determine_snp_summary(tbl_calling_results)
    ws = wb.create_sheet(title='SNP summary')
    for row in tbl_snp_summary.cutout('ASSAY_ID'):
        ws.append(row)
    
    tbl_concordance_summary = determine_concordance_summary(tbl_calling_results)
    ws = wb.create_sheet(title='Concordance summary')
    for row in tbl_concordance_summary:
        ws.append(row)
    
    tbl_snp_concordance_summary = determine_snp_concordance_summary(tbl_calling_results)
    ws = wb.create_sheet(title='SNP concordance')
    for row in tbl_snp_concordance_summary:
        ws.append(row)
    
    tbl_sample_concordance_summary = determine_sample_concordance_summary(tbl_calling_results)
    ws = wb.create_sheet(title='Sample concordance')
    for row in tbl_sample_concordance_summary:
        ws.append(row)
    
    tbl_concordance_matrix = determine_concordance_matrix(tbl_calling_results)
    ws = wb.create_sheet(title='concordance matrix')
    for row in tbl_concordance_matrix:
        ws.append(row)
        
    assay_ids = (tbl_snp_summary
        .values('ASSAY_ID')
        .array()
    )
    for assay_id in assay_ids:
        tbl_concordance_matrix = determine_concordance_matrix(tbl_calling_results, assay_id)
        chrom = tbl_snp_summary.selecteq('ASSAY_ID', assay_id).values('Chrom')[0]
        pos = tbl_snp_summary.selecteq('ASSAY_ID', assay_id).values('Pos')[0]
        ws = wb.create_sheet(title='conc %s %s' % (chrom, pos))
        for row in tbl_concordance_matrix:
            ws.append(row)

    
    wb.save(summary_fn)

create_summary_tables(plate_name, calling_parameters_fn, sample_mappings_fn, ref_genome_fn, illumina_vcf_fn)

# <codecell>

def rerun_calling(plate_name, calling_parameters_fn, sample_mappings_fn, ref_genome_fn, illumina_vcf_fn,
              recon_id_name='sample_code'):
    tbl_sample_mappings = cache_sample_mappings(sample_mappings_fn, plate_name, recon_id_name=recon_id_name, rewrite=False)
    tbl_sanger_sequenom = cache_raw_sequenom_data(plate_name, tbl_sample_mappings, ref_genome_fn, rewrite=False)
    tbl_illumina = cache_illumina(illumina_vcf_fn, tbl_sanger_sequenom, plate_name, rewrite=False)
    tbl_calling_parameters = cache_calling_parameters(calling_parameters_fn, plate_name, rewrite=True)
    tbl_calling_results = cache_calling_results(tbl_sanger_sequenom, tbl_calling_parameters, tbl_illumina,
                                                plate_name, ref_genome_fn, rewrite=True)
    return(tbl_calling_results)

# <codecell>

def create_all_plots(tbl_calling_results, tbl_calling_parameters, plate_name, low_intensity_plot_scale=3):
    for genotype_column in ['called_genotype', 'called_gt', 'illumina_genotype', 'illumina_gt',
                            'SEQUENOM_GENOTYPE', 'sequenom_gt']:
        plot_fn = "%s/plots_%s_%s.pdf" % (plot_dir_format % (data_dir, plate_name), plate_name, genotype_column)
        if not os.path.exists(os.path.dirname(plot_fn)):
            os.makedirs(os.path.dirname(plot_fn))
        with PdfPages(plot_fn) as pdf:
            for plot_type in ['height', 'theta']:
                for plot_scale in [0, low_intensity_plot_scale]:
                    print(genotype_column, plot_type, plot_scale)
                    sequenom_scatter_plot(tbl_calling_results, tbl_calling_parameters, genotype_column,
                                          plot_type=plot_type, plot_scale=plot_scale)
                    pdf.savefig()
    traffic_light_plot(tbl_calling_results)
    plot_fn = "%s/traffic_light_plots_%s.pdf" % (plot_dir_format % (data_dir, plate_name), plate_name)
    savefig(plot_fn)

# <codecell>

def calling_pipeline(plate_name, calling_parameters_fn, sample_mappings_fn, ref_genome_fn, illumina_vcf_fn,
                     recon_id_name='sample_code', rewrite=False):
    create_all_plots(
        load_data(plate_name, calling_parameters_fn, sample_mappings_fn, ref_genome_fn, illumina_vcf_fn,
                  recon_id_name=recon_id_name, rewrite=rewrite),
        cache_calling_parameters(calling_parameters_fn, plate_name, rewrite=rewrite),
        plate_name
    )
    create_summary_tables(plate_name, calling_parameters_fn, sample_mappings_fn, ref_genome_fn, illumina_vcf_fn)

# <codecell>

def rerun_calling_pipeline(plate_name, calling_parameters_fn, sample_mappings_fn, ref_genome_fn, illumina_vcf_fn,
                           recon_id_name='sample_code'):
    create_all_plots(
        rerun_calling(plate_name, calling_parameters_fn, sample_mappings_fn, ref_genome_fn, illumina_vcf_fn,
                      recon_id_name=recon_id_name),
        cache_calling_parameters(calling_parameters_fn, plate_name, rewrite=True),
        plate_name
    )
    create_summary_tables(plate_name, calling_parameters_fn, sample_mappings_fn, ref_genome_fn, illumina_vcf_fn)

# <headingcell level=1>

# Run full pipeline

# <codecell>

# Initial test plate with 384 samples
calling_pipeline(plate_name, calling_parameters_fn, sample_mappings_fn, ref_genome_fn, illumina_vcf_fn, rewrite=True)

# <codecell>

# Rerun initial test plate with new parameters
rerun_calling_pipeline(
    plate_name,
    '/nfs/team112_internal/rp7/recon/sanger_sequenom/parameters/W36000_QC409376_409377_409378_409379_20150715/W36000_QC409376_409377_409378_409379_20150715_richard_noillumina.txt',
    sample_mappings_fn, ref_genome_fn, illumina_vcf_fn
)

# <codecell>

# First sprint plate with 30 actual samples and 66 blanks, using parameters from test plate
calling_pipeline(
    'W36000_Replication411601____20150728',
    '/nfs/team112_internal/rp7/recon/sanger_sequenom/parameters/W36000_QC409376_409377_409378_409379_20150715/W36000_QC409376_409377_409378_409379_20150715_richard_noillumina.txt',
    '/nfs/team112_internal/rp7/recon/sanger_sequenom/sample_mappings/3732stdy_manifest_4039_270715.xls',
    ref_genome_fn,
    illumina_vcf_fn,
    recon_id_name='SUPPLIER SAMPLE NAME',
    rewrite=True
)

# <codecell>

# First sprint plate with 30 actual samples and 66 blanks, using optimised parameters
rerun_calling_pipeline(
    'W36000_Replication411601____20150728',
    '/nfs/team112_internal/rp7/recon/sanger_sequenom/parameters/W36000_Replication411601____20150728/W36000_Replication411601____20150728_richard_noillumina.txt',
    '/nfs/team112_internal/rp7/recon/sanger_sequenom/sample_mappings/3732stdy_manifest_4039_270715.xls',
    ref_genome_fn,
    illumina_vcf_fn,
    recon_id_name='SUPPLIER SAMPLE NAME'
)

# <codecell>


# <codecell>


# <headingcell level=1>

# Summary tables

# <codecell>

tbl_calling_results = load_data(plate_name, calling_parameters_fn, sample_mappings_fn, ref_genome_fn, illumina_vcf_fn)
tbl_calling_results

# <codecell>


# <codecell>


# <codecell>


# <codecell>

results_summary = collections.OrderedDict()
results_summary['a) # wells'] = len(tbl_calling_results.distinct('WELL_POSITION').data())
results_summary['b) # samples'] = len(tbl_calling_results.distinct('WELL_POSITION').selectnotin('recon_id', water_names).data())
results_summary['c) # sequenced samples'] = len(tbl_calling_results.distinct('WELL_POSITION').selectne('illumina_gt', '?').data())
results_summary['d) # waters'] = len(tbl_calling_results.distinct('WELL_POSITION').selectin('recon_id', water_names).data())
results_summary['e) # Pf calls samples'] = len(
    tbl_calling_results
    .selecteq('species', 'Pf')
    .selectnotin('recon_id', water_names)
    .selectin('called_gt', ['0/0', '0/1', '1/1'])
    .data()
)
results_summary['f) # Pf missing calls samples'] = len(
    tbl_calling_results
    .selecteq('species', 'Pf')
    .selectnotin('recon_id', water_names)
    .selectnotin('called_gt', ['0/0', '0/1', '1/1'])
    .data()
)
results_summary['g) # Pf calls water'] = len(
    tbl_calling_results
    .selecteq('species', 'Pf')
    .selectin('recon_id', water_names)
    .selectin('called_gt', ['0/0', '0/1', '1/1'])
    .data()
)
results_summary['h) # Pf missing calls water'] = len(
    tbl_calling_results
    .selecteq('species', 'Pf')
    .selectin('recon_id', water_names)
    .selectnotin('called_gt', ['0/0', '0/1', '1/1'])
    .data()
)
results_summary['i) % samples called'] = round(100.0*(results_summary['e) # Pf calls samples'] /
                            (results_summary['e) # Pf calls samples'] + results_summary['f) # Pf missing calls samples'])), 2)
results_summary['j) % water missing'] = round(100.0*(results_summary['h) # Pf missing calls water'] /
                            (results_summary['h) # Pf missing calls water'] + results_summary['g) # Pf calls water'])), 2)

print(results_summary)
etl.fromdicts([results_summary]).transpose().pushheader(['variable', 'value']).displayall()

# <codecell>

sample_aggregation = collections.OrderedDict()
sample_aggregation['# calls'] = 'called_gt', lambda rec: np.sum(x in ['0/0', '0/1', '1/1'] for x in list(rec))
# aggregation['# calls'] = len

# aggregation['#GATK'] = 'IsInGATK', sum
# aggregation['#GATK accessible'] = ('IsInGATK', 'IsAccessible'), lambda rec: np.sum([x[0] and x[1] for x in list(rec)])
# aggregation['#TP'] = ('IsInGATK', 'IsInTruth'), lambda rec: np.sum([x[0] and x[1] for x in list(rec)])
# aggregation['#FP'] = ('IsInGATK', 'IsInTruth', 'IsAccessible'), lambda rec: np.sum([x[0] and not x[1] and x[2] for x in list(rec)])
# aggregation['#FN'] = ('IsInGATK', 'IsInTruth'), lambda rec: np.sum([not x[0] and x[1] for x in list(rec)])
# aggregation['#TPmod3'] = ('REF', 'ALT', 'IsInGATK', 'IsInTruth'), lambda rec: np.sum([(len(x[0]) - len(x[1])) % 3 == 0 and x[2] and x[3] for x in list(rec)])
# aggregation['#FPmod3'] = ('REF', 'ALT', 'IsInGATK', 'IsInTruth', 'IsAccessible'), lambda rec: np.sum([(len(x[0]) - len(x[1])) % 3 == 0 and x[2] and not x[3] and x[4] for x in list(rec)])
# aggregation['#FNmod3'] = ('REF', 'ALT', 'IsInGATK', 'IsInTruth'), lambda rec: np.sum([(len(x[0]) - len(x[1])) % 3 == 0 and not x[2] and x[3] for x in list(rec)])

tbl_calls_per_sample = (tbl_calling_results
    .selecteq('species', 'Pf')
    .selectnotin('recon_id', water_names)
    .aggregate('recon_id', sample_aggregation)
)
tbl_calls_per_sample.valuecounts('# calls').sort('# calls', reverse=True).displayall()
tbl_calls_per_sample.selectlt('# calls', 10).sort('# calls', reverse=False).displayall()

# <codecell>

cache_sample_mappings(sample_mappings_fn, plate_name).valuecounts('recon_id').displayall()

# <codecell>

well_aggregation = collections.OrderedDict()
well_aggregation['# calls'] = 'called_gt', lambda rec: np.sum(x in ['0/0', '0/1', '1/1'] for x in list(rec))
well_aggregation['recon_id'] = 'recon_id', lambda rec: list(rec)[0]

tbl_calls_per_well = (tbl_calling_results
    .selecteq('species', 'Pf')
#     .selectnotin('recon_id', water_names)
    .aggregate('WELL_POSITION', well_aggregation)
#     .leftjoin(tbl_calling_results)
    .addfield('row_number', lambda rec: ord(rec['WELL_POSITION'][0]) - 64)
    .addfield('column_number', lambda rec: int(rec['WELL_POSITION'][1:3]))
    .addfield('traffic_light', determine_traffic_light)
)

# <codecell>

traffic_light_plot()
plot_fn = "%s/traffic_light_plots_%s.pdf" % (plot_dir_format % (data_dir, plate_name), plate_name)
savefig(plot_fn)

# <codecell>

tbl_calls_per_well

# <codecell>

tbl_calls_per_well.valuecounts('traffic_light')

# <codecell>

def

# <codecell>

tbl_calling_results.valuecounts()

# <codecell>

tbl_calls_per_well.valuecounts('# calls').displayall()

# <codecell>

tbl_calls_per_well.displayall()

# <codecell>

(etl
                .fromxlsx(sample_mappings_fn, data_only=True)).selecteq('sample_code', 'PA0011-C')

# <codecell>

snp_aggregation = collections.OrderedDict()
snp_aggregation['Chrom'] = 'chrom', lambda rec: list(rec)[0]
snp_aggregation['Pos'] = 'pos', lambda rec: list(rec)[0]
snp_aggregation['# Ref'] = ('recon_id', 'called_gt'), lambda rec: np.sum(not(x[0] in water_names) and x[1] == '0/0' for x in list(rec))
snp_aggregation['# Het'] = ('recon_id', 'called_gt'), lambda rec: np.sum(not(x[0] in water_names) and x[1] == '0/1' for x in list(rec))
snp_aggregation['# Alt'] = ('recon_id', 'called_gt'), lambda rec: np.sum(not(x[0] in water_names) and x[1] == '1/1' for x in list(rec))
snp_aggregation['# No-call'] = ('recon_id', 'called_gt'), lambda rec: np.sum(not(x[0] in water_names) and x[1] == './.' for x in list(rec))
snp_aggregation['# Water Ref'] = ('recon_id', 'called_gt'), lambda rec: np.sum((x[0] in water_names) and x[1] == '0/0' for x in list(rec))
snp_aggregation['# Water Het'] = ('recon_id', 'called_gt'), lambda rec: np.sum((x[0] in water_names) and x[1] == '0/1' for x in list(rec))
snp_aggregation['# Water Alt'] = ('recon_id', 'called_gt'), lambda rec: np.sum((x[0] in water_names) and x[1] == '1/1' for x in list(rec))
snp_aggregation['# Water No-call'] = ('recon_id', 'called_gt'), lambda rec: np.sum((x[0] in water_names) and x[1] == './.' for x in list(rec))

tbl_snp_summary = (tbl_calling_results
    .selecteq('species', 'Pf')
    .aggregate('ASSAY_ID', snp_aggregation)
)

tbl_snp_summary.cutout('ASSAY_ID').displayall()

# <codecell>

tbl_calling_results.valuecounts('recon_id')

# <codecell>

(tbl_calling_results
    .selecteq('species', 'Pf')
    .selectnotin('recon_id', water_names)).valuecounts('called_gt')

# <codecell>

(tbl_calling_results
    .selecteq('species', 'Pf')
    .selectne('illumina_gt', '?')
).valuecounts('called_gt').replaceall(None, 0)

# <codecell>

(tbl_calling_results
    .selecteq('species', 'Pf')
    .selectne('illumina_gt', '?')
).pivot('called_gt', 'illumina_gt', 'called_gt', len).replaceall(None, 0)

# <codecell>

assay_ids = (tbl_calling_results
    .selecteq('species', 'Pf')
    .distinct('ASSAY_ID')
    .values('ASSAY_ID')
    .array()
)
for assay_id in assay_ids:
    print(assay_id)
    (tbl_calling_results
        .selecteq('ASSAY_ID', assay_id)
        .selectne('illumina_gt', '?')
        .pivot('called_gt', 'illumina_gt', 'called_gt', len)
        .replaceall(None, 0)
        .displayall()
    )
    

# <codecell>

snp_concordance_aggregation = collections.OrderedDict()
snp_concordance_aggregation['Chrom'] = 'chrom', lambda rec: list(rec)[0]
snp_concordance_aggregation['Pos'] = 'pos', lambda rec: list(rec)[0]
snp_concordance_aggregation['# Assays'] = len
snp_concordance_aggregation['# Concordant'] = ('called_gt', 'illumina_gt'), lambda rec: np.sum(x[0] != './.' and x[0] == x[1] for x in list(rec))
snp_concordance_aggregation['# Discordant'] = ('called_gt', 'illumina_gt'), lambda rec: np.sum(x[0] != './.' and x[1] != './.' and x[0] != x[1] for x in list(rec))
snp_concordance_aggregation['# Hom discordant'] = ('called_gt', 'illumina_gt'), lambda rec: np.sum((x[0] == '0/0' or x[0] == '1/1') and (x[1] == '0/0' or x[1] == '1/1') and x[0] != x[1] for x in list(rec))
snp_concordance_aggregation['# Het discordant'] = ('called_gt', 'illumina_gt'), lambda rec: np.sum((x[0] == '0/0' and x[1] == '0/1') or (x[0] == '1/1' and x[1] == '0/1') or (x[0] == '0/1' and x[1] == '0/0') or (x[0] == '0/1' and x[1] == '1/1') for x in list(rec))
snp_concordance_aggregation['# No call Sequenom'] = ('called_gt', 'illumina_gt'), lambda rec: np.sum(x[0] == './.' and x[1] != './.' for x in list(rec))
snp_concordance_aggregation['# No call Illumina'] = ('called_gt', 'illumina_gt'), lambda rec: np.sum(x[0] != './.' and x[1] == './.' for x in list(rec))
snp_concordance_aggregation['# No call both'] = ('called_gt', 'illumina_gt'), lambda rec: np.sum(x[0] == './.' and x[1] == './.' for x in list(rec))

tbl_snp_concordance_summary = (tbl_calling_results
    .selecteq('species', 'Pf')
    .selectne('illumina_gt', '?')
    .aggregate('ASSAY_ID', snp_concordance_aggregation)
    .addfield('% Concordance', lambda rec: round((rec['# Concordant'] / (rec['# Concordant'] + rec['# Discordant'])) * 100, 2))
    .addfield('% Hom discordance', lambda rec: round((rec['# Hom discordant'] / (rec['# Concordant'] + rec['# Discordant'])) * 100, 2))
    .addfield('% Missing either assay', lambda rec: round(((rec['# No call Sequenom'] + rec['# No call Illumina'] + rec['# No call both']) / rec['# Assays']) * 100, 2))
)

tbl_snp_concordance_summary.cutout('ASSAY_ID').displayall()

# <codecell>

concordance_aggregation = collections.OrderedDict()
concordance_aggregation['# Assays'] = len
concordance_aggregation['# Concordant'] = ('called_gt', 'illumina_gt'), lambda rec: np.sum(x[0] != './.' and x[0] == x[1] for x in list(rec))
concordance_aggregation['# Discordant'] = ('called_gt', 'illumina_gt'), lambda rec: np.sum(x[0] != './.' and x[1] != './.' and x[0] != x[1] for x in list(rec))
concordance_aggregation['# Hom discordant'] = ('called_gt', 'illumina_gt'), lambda rec: np.sum((x[0] == '0/0' or x[0] == '1/1') and (x[1] == '0/0' or x[1] == '1/1') and x[0] != x[1] for x in list(rec))
concordance_aggregation['# Het discordant'] = ('called_gt', 'illumina_gt'), lambda rec: np.sum((x[0] == '0/0' and x[1] == '0/1') or (x[0] == '1/1' and x[1] == '0/1') or (x[0] == '0/1' and x[1] == '0/0') or (x[0] == '0/1' and x[1] == '1/1') for x in list(rec))
concordance_aggregation['# No call Sequenom'] = ('called_gt', 'illumina_gt'), lambda rec: np.sum(x[0] == './.' and x[1] != './.' for x in list(rec))
concordance_aggregation['# No call Illumina'] = ('called_gt', 'illumina_gt'), lambda rec: np.sum(x[0] != './.' and x[1] == './.' for x in list(rec))
concordance_aggregation['# No call both'] = ('called_gt', 'illumina_gt'), lambda rec: np.sum(x[0] == './.' and x[1] == './.' for x in list(rec))

tbl_concordance_summary = (tbl_calling_results
    .selecteq('species', 'Pf')
    .selectne('illumina_gt', '?')
    .aggregate('plate_code', concordance_aggregation)
    .addfield('% Concordance', lambda rec: round((rec['# Concordant'] / (rec['# Concordant'] + rec['# Discordant'])) * 100, 2))
    .addfield('% Hom discordance', lambda rec: round((rec['# Hom discordant'] / (rec['# Concordant'] + rec['# Discordant'])) * 100, 2))
    .addfield('% Missing either assay', lambda rec: round(((rec['# No call Sequenom'] + rec['# No call Illumina'] + rec['# No call both']) / rec['# Assays']) * 100, 2))
)

tbl_concordance_summary.transpose().displayall()

# <codecell>

tbl_sample_concordance_summary = (tbl_calling_results
    .selecteq('species', 'Pf')
    .selectne('illumina_gt', '?')
    .aggregate('recon_id', concordance_aggregation)
    .addfield('% Concordance', lambda rec: round((rec['# Concordant'] / (rec['# Concordant'] + rec['# Discordant'])) * 100, 2))
    .addfield('% Hom discordance', lambda rec: round((rec['# Hom discordant'] / (rec['# Concordant'] + rec['# Discordant'])) * 100, 2))
    .addfield('% Missing either assay', lambda rec: round(((rec['# No call Sequenom'] + rec['# No call Illumina'] + rec['# No call both']) / rec['# Assays']) * 100, 2))
)

(tbl_sample_concordance_summary
    .selectne('# Concordant', 10)
    .sort('# No call Illumina', reverse=True)
    .sort('# No call Sequenom', reverse=True)
    .sort('% Concordance')
    .sort('% Hom discordance', reverse=True)
    .displayall()
)

# <headingcell level=1>

# Sprint exploration

# <codecell>

tbl_sprint = load_data(
    'W36000_Replication411601____20150728',
    '/nfs/team112_internal/rp7/recon/sanger_sequenom/sample_mappings/3732stdy_manifest_4039_270715.xls',
    ref_genome_fn,
    illumina_vcf_fn,
    recon_id_name='SUPPLIER SAMPLE NAME'
)

# <codecell>

tbl_sprint

# <codecell>

tbl_sprint.valuecounts('illumina_gt')

# <codecell>

tbl_sprint.valuecounts('recon_id').displayall()

# <headingcell level=1>

# Load data

# <codecell>

tbl_sample_mappings = cache_sample_mappings(sample_mappings_fn, plate_name, rewrite=True)
# tbl_sample_mappings

# <codecell>

tbl_sanger_sequenom = cache_raw_sequenom_data(plate_name, tbl_sample_mappings, ref_genome_fn, rewrite=True)
# tbl_sanger_sequenom.distinct('ASSAY_ID').values('ASSAY_ID').array()
# tbl_sanger_sequenom

# <codecell>

tbl_illumina = cache_illumina(illumina_vcf_fn, tbl_sanger_sequenom, plate_name, rewrite=True)

# <codecell>

tbl_calling_parameters = cache_calling_parameters(calling_parameters_fn, plate_name, rewrite=True)
# tbl_calling_parameters.displayall()

# <codecell>

tbl_calling_results = cache_calling_results(tbl_sanger_sequenom, tbl_calling_parameters, tbl_illumina,
                                            plate_name, ref_genome_fn, rewrite=True)
# len(tbl_calling_results)

# <codecell>


# <headingcell level=1>

# Scatter plots

# <codecell>

for genotype_column in ['SEQUENOM_GENOTYPE', 'sequenom_gt', 'called_genotype', 'called_gt', 'illumina_genotype', 'illumina_gt']:
    plot_fn = "%s/plots_%s.pdf" % (plot_dir_format % (data_dir, plate_name), genotype_column)
    if not os.path.exists(os.path.dirname(plot_fn)):
        os.makedirs(os.path.dirname(plot_fn))
    with PdfPages(plot_fn) as pdf:
        for plot_type in ['height', 'theta']:
            for plot_scale in ['full', 'low_intensity']:
                print(genotype_column, plot_type, plot_scale)
                sequenom_scatter_plot(tbl_calling_results, genotype_column, plot_type=plot_type, plot_scale=plot_scale)
                pdf.savefig()

# <codecell>


# <headingcell level=1>

# Explore discordances between Sequenom and Illumina

# <codecell>

tbl_calling_results.selectne('illumina_gt', '?').valuecounts('sequenom_gt', 'illumina_gt').displayall()

# <headingcell level=2>

# Homozygote discordance

# <codecell>

tbl_calling_results.selecteq('sequenom_gt', '0/0').selecteq('illumina_gt', '1/1')

# <codecell>

tbl_calling_results.selecteq('recon_id', 'PN0022-C').displayall()

# <codecell>

tbl_calling_results.selecteq('sequenom_gt', '1/1').selecteq('illumina_gt', '0/0')

# <codecell>

tbl_calling_results.selecteq('recon_id', 'PH0029-Cx').displayall()

# <headingcell level=2>

# Heterozygote discordance

# <codecell>

tbl_calling_results.selecteq('sequenom_gt', '0/1').selecteq('illumina_gt', '0/0')

# <codecell>

tbl_calling_results.selecteq('sequenom_gt', '0/0').selecteq('illumina_gt', '0/1').selecteq('recon_id', 'PF0077-C').displayall()

# <codecell>

tbl_calling_results.selecteq('sequenom_gt', '0/0').selecteq('illumina_gt', '0/1').selectne('recon_id', 'PF0077-C').displayall()

# <codecell>

tbl_calling_results.selecteq('sequenom_gt', '1/1').selecteq('illumina_gt', '0/1').displayall()

# <headingcell level=2>

# Missing Illumina calls

# <codecell>

tbl_calling_results.selecteq('illumina_gt', './.').selecteq('AD', [0,0]).displayall()

# <headingcell level=2>

# C580Y

# <codecell>

tbl_calling_results.selecteq('pos', 1725259).selecteq('illumina_gt', '0/1')

# <codecell>


