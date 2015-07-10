# -*- coding: utf-8 -*-
# <nbformat>3.0</nbformat>

# <headingcell level=1>

# Setup

# <codecell>

%run _shared_setup.ipynb

import openpyxl
import locale
from Bio import SeqIO
import petlx.bio

# <codecell>

GATC_SEQ_DIRs = collections.OrderedDict()
GATC_SEQ_DIRs['Forward'] = '/data/plasmodium/pfalciparum/recon/data/raw/capillary/GATC/1181127'
GATC_SEQ_DIRs['Reverse'] = '/data/plasmodium/pfalciparum/recon/data/raw/capillary/GATC/1181139'
GATC_SEQ_DIRs['ForwardRep'] = '/data/plasmodium/pfalciparum/recon/data/raw/capillary/GATC/1210119'

CAPILLARY_SAMPLES_FN = '/data/plasmodium/pfalciparum/recon/data/processed/capillary/GATC_less_DNA.xlsx'
!mkdir -p /data/plasmodium/pfalciparum/recon/data/processed/capillary

BAM_FILES_DIR = '/data/plasmodium/pfalciparum/recon/data/processed/capillary/bams/GATC_less_DNA'
GATC_BAM_FILES_FOFN = '/data/plasmodium/pfalciparum/recon/data/processed/capillary/bams/GATC_less_DNA/gatc_bams.%s.list'
!mkdir -p {BAM_FILES_DIR}

VARIANT_FILES_DIR = '/data/plasmodium/pfalciparum/recon/data/processed/capillary/variants/GATC_less_DNA'
!mkdir -p {VARIANT_FILES_DIR}
GATC_CAPILLARY_VCF_FN = "%s/gatc_capillary.%%s.vcf" % VARIANT_FILES_DIR
ILLUMINA_CALLS_EXCEL_FN = "%s/illumina_calls.xlsx" % VARIANT_FILES_DIR
SEQUENOM_CALLS_EXCEL_FN = "%s/sequenom_calls.xlsx" % VARIANT_FILES_DIR
CAPILLARY_CALLS_EXCEL_FN = "%s/capillary_calls.xlsx" % VARIANT_FILES_DIR
MERGED_CALLS_EXCEL_FN = "%s/merged_calls.xlsx" % VARIANT_FILES_DIR

SQNM_VCF_FN = DATA_DIR + '/Sqnm_data_DK1066_W1378_20150603.vcf.gz'
SAMPLES_FN = DATA_DIR + '/%s_capillary_samples_less_DNA.txt'

PLOT_DIR = '/Users/rpearson/Documents/projects/Recon/Slides/20150710_assess_GATC_less_DNA'
!mkdir -p {PLOT_DIR}

# <codecell>

!mkdir -p {GATC_SEQ_DIRs['ForwardRep']}
!cp /Users/rpearson/Dropbox/69d4b150f584e230336d8f66ba06851aafa903d4/WatchBox_1210119/* {GATC_SEQ_DIRs['ForwardRep']}/

# <codecell>

SAMPLES_FN

# <headingcell level=1>

# Parse sample details

# <codecell>

def convert_to_fastq(seq_fn='/data/plasmodium/pfalciparum/recon/data/raw/capillary/GATC/1181127/D07_PD0004-C-12891.fas',
                     fastq_fn=None, bad_qual_distance=None, qual_good=60, qual_bad=0):
    if fastq_fn is None:
        fastq_fn = seq_fn.replace('.fas', '.%s.fastq' % bad_qual_distance).replace('.seq', '.%s.fastq' % bad_qual_distance)
        
    # make fastq
    with open(seq_fn, "r") as fasta, open(fastq_fn, "w") as fastq:
        for record in SeqIO.parse(fasta, "fasta"):
            record.seq = record.seq.tomutable()
            record.letter_annotations["phred_quality"] = [qual_good] * len(record)
            if bad_qual_distance is not None:
                if bad_qual_distance==0:
                    for i in range(len(record)):
                        if record.seq[i].islower() or record.seq[i].upper()=='N':
                            record.letter_annotations["phred_quality"][i] = qual_bad
                elif bad_qual_distance==1:
                    for i in range(len(record)):
                        if i == 0:
                            if record.seq[i].islower() or record.seq[i+1].islower() or record.seq[i].upper()=='N':
                                record.letter_annotations["phred_quality"][i] = qual_bad
                        elif i == (len(record) - 1):
                            if record.seq[i].islower() or record.seq[i-1].islower() or record.seq[i].upper()=='N':
                                record.letter_annotations["phred_quality"][i] = qual_bad
                        else:
                            if record.seq[i].islower() or record.seq[i-1].islower() or record.seq[i+1].islower() or record.seq[i].upper()=='N':
                                record.letter_annotations["phred_quality"][i] = qual_bad
                        # Change N's to a, as bwa seems to ignore these and treat as deletion. These should have qual=0 and hence not get called as SNPs
                        if record.seq[i].upper()=='N':
                            record.seq[i] = 'a'
                elif bad_qual_distance==2:
                    for i in range(len(record)):
                        if i == 0:
                            if record.seq[i].islower() or record.seq[i+1].islower() or record.seq[i+2].islower() or record.seq[i].upper()=='N':
                                record.letter_annotations["phred_quality"][i] = qual_bad
                        if i == 1:
                            if record.seq[i].islower() or record.seq[i+1].islower() or record.seq[i-1].islower() or record.seq[i].upper()=='N':
                                record.letter_annotations["phred_quality"][i] = qual_bad
                        elif i == (len(record) - 1):
                            if record.seq[i].islower() or record.seq[i-1].islower() or record.seq[i-2].islower() or record.seq[i].upper()=='N':
                                record.letter_annotations["phred_quality"][i] = qual_bad
                        elif i == (len(record) - 2):
                            if record.seq[i].islower() or record.seq[i-1].islower() or record.seq[i+1].islower() or record.seq[i].upper()=='N':
                                record.letter_annotations["phred_quality"][i] = qual_bad
                        else:
                            if record.seq[i].islower() or record.seq[i-1].islower() or record.seq[i+1].islower() or record.seq[i-2].islower() or record.seq[i+2].islower() or record.seq[i].upper()=='N':
                                record.letter_annotations["phred_quality"][i] = qual_bad
                        # Change N's to a, as bwa seems to ignore these and treat as deletion. These should have qual=0 and hence not get called as SNPs
                        if record.seq[i].upper()=='N':
                            record.seq[i] = 'a'
                else:
                    print("Warning - can't handle bad_qual_distance=%s, assuming bad_qual_distance=None" % bad_qual_distance)
                        
            SeqIO.write(record, fastq, "fastq")
    
    return(fastq_fn)

convert_to_fastq()

# <codecell>

convert_to_fastq(bad_qual_distance=0)

# <codecell>

convert_to_fastq(bad_qual_distance=1)

# <codecell>

convert_to_fastq(bad_qual_distance=2)

# <codecell>

def sequence_length(seq_fn='/data/plasmodium/pfalciparum/recon/data/raw/capillary/GATC/1181127/D07_PD0004-C-12891.fas'):
    return(len(SeqIO.read(open(seq_fn, "r"), "fasta").seq))
sequence_length()

# <codecell>

forward_fas_files = !ls {GATC_SEQ_DIRs['Forward']}/*.fas
reverse_fas_files = !ls {GATC_SEQ_DIRs['Reverse']}/*.fas
forward_rep_fas_files = !ls {GATC_SEQ_DIRs['ForwardRep']}/*.fas

# <codecell>

tbl_gatc_metadata = (etl.wrap([[x] for x in forward_fas_files])
    .pushheader(['seq_fn'])
    .addfield('ox_code', lambda rec: rec[0].replace(GATC_SEQ_DIRs['Forward'], '').replace('-12891.fas', '')[5:])
    .replace('ox_code', 'PD0005-1', 'PD0005-01')
    .replace('ox_code', 'PH0883-C', 'PH0883-Cx')
    .addfield('primers', 'Forward')
    .addfield('seq_length', lambda rec: sequence_length(rec['seq_fn']))
).cat(
(etl.wrap([[x] for x in reverse_fas_files])
    .pushheader(['seq_fn'])
    .addfield('ox_code', lambda rec: rec[0].replace(GATC_SEQ_DIRs['Reverse'], '').replace('-12825seq2.fas', '')[5:])
    .replace('ox_code', 'PD0005-1', 'PD0005-01')
    .replace('ox_code', 'PH0883-C', 'PH0883-Cx')
    .addfield('primers', 'Reverse')
    .addfield('seq_length', lambda rec: sequence_length(rec['seq_fn']))
)).cat(
(etl.wrap([[x] for x in forward_rep_fas_files])
    .pushheader(['seq_fn'])
    .addfield('ox_code', lambda rec: rec[0].replace(GATC_SEQ_DIRs['ForwardRep'], '').replace('-12891.fas', '')[5:])
    .replace('ox_code', 'PD0005-1', 'PD0005-01')
    .replace('ox_code', 'PH0883-C', 'PH0883-Cx')
    .addfield('primers', 'ForwardRep')
    .addfield('seq_length', lambda rec: sequence_length(rec['seq_fn']))
)).sort('ox_code')
print(len(tbl_gatc_metadata.data()))
tbl_gatc_metadata.displayall(index_header=True)

# <codecell>

vcf_reader = vcf.Reader(filename=PGV4_VCF_FN)
pgv4_samples = etl.wrap([[x] for x in vcf_reader.samples]).pushheader(['ox_code']).addfield('is_in_pgv4_vcf', True)

# <codecell>

tbl_sample_sets

# <codecell>

tbl_pgv4_sample_manifest

# <codecell>

pgv4_samples

# <codecell>

tbl_capillary_samples = (
    tbl_gatc_metadata
    .cut(['ox_code'])
    .distinct('ox_code')
    .leftjoin(tbl_sample_sets, lkey='ox_code', rkey='sample_code')
    .addfield('have_sequenom_data', lambda rec: rec['source_code'] is not None)
    .cut(['ox_code', 'have_sequenom_data'])
    .leftjoin(pgv4_samples)
    .leftjoin(tbl_pgv4_sample_manifest.selecteq('Exclude', 0), lkey='ox_code', rkey='Sample')
    .leftjoin(tbl_pgv4_sample_manifest, lkey='ox_code', rkey='Sample')
#     .cut(['ox_code', 'company', 'have_sequenom_data', 'is_in_pgv4_vcf', 'Pf3k v3'])
    .cut(['ox_code', 'have_sequenom_data', 'is_in_pgv4_vcf', 'Fws'])
    .replace('have_sequenom_data', False, None)
#     .convert('Pf3k v3', bool)
#     .replace('Pf3k v3', False, None)
    .leftjoin(tbl_gatc_metadata.selecteq('primers', 'Forward').cut(['ox_code', 'seq_length']), lkey='ox_code', rkey='ox_code')
    .rename('seq_length', 'forward')
    .leftjoin(tbl_gatc_metadata.selecteq('primers', 'Reverse').cut(['ox_code', 'seq_length']), lkey='ox_code', rkey='ox_code')
    .rename('seq_length', 'reverse')
    .leftjoin(tbl_gatc_metadata.selecteq('primers', 'ForwardRep').cut(['ox_code', 'seq_length']), lkey='ox_code', rkey='ox_code')
    .rename('seq_length', 'forward_rep')
    .selectne('ox_code', 'WATER')
    .sort(['ox_code'])
)
print(len(tbl_capillary_samples.data()))
tbl_capillary_samples.displayall()

# <codecell>

tbl_capillary_samples.selecteq('forward_rep', 0).cut('ox_code').displayall()

# <codecell>

np.min(tbl_capillary_samples.selectnotnone('Fws').values('Fws').array())

# <codecell>

# sns.jointplot(tbl_capillary_samples.values('forward').array(), tbl_capillary_samples.values('reverse').array())
sns.jointplot('forward', 'reverse', tbl_capillary_samples.todataframe())
plt.savefig('%s/forward_vs_reverse.pdf' % PLOT_DIR, format='pdf')

# <codecell>

# sns.jointplot(tbl_capillary_samples.values('forward').array(), tbl_capillary_samples.values('reverse').array())
sns.jointplot('forward', 'forward_rep', tbl_capillary_samples.todataframe())
plt.savefig('%s/forward_vs_rep.pdf' % PLOT_DIR, format='pdf')

# <codecell>

# sns.jointplot(tbl_capillary_samples.values('forward').array(), tbl_capillary_samples.values('reverse').array())
sns.distplot(
    tbl_capillary_samples.addfield('Total length', lambda rec: rec['forward'] + rec['reverse']).values('Total length').array(),
    kde=False,
    bins=arange(200, 2000, 100),
    axlabel='Total length'
)
plt.savefig('%s/total_length_dist.pdf' % PLOT_DIR, format='pdf')

# <codecell>

# sns.jointplot(tbl_capillary_samples.values('forward').array(), tbl_capillary_samples.values('reverse').array())
sns.distplot(
    tbl_capillary_samples.addfield('Total length', lambda rec: rec['forward_rep'] + rec['reverse']).values('Total length').array(),
    kde=False,
    bins=arange(200, 2000, 100),
    axlabel='Total length'
)
plt.savefig('%s/total_length_dist_rep.pdf' % PLOT_DIR, format='pdf')

# <codecell>

tbl_capillary_samples.addfield('Covers full sequence', lambda rec: rec['forward_rep'] + rec['reverse'] >= 1012).valuecounts('Covers full sequence')

# <codecell>

tbl_capillary_samples.toxlsx(CAPILLARY_SAMPLES_FN)

# <codecell>

def run_bwa(ox_code='PA0155-C', company='GATC', primers='Forward',
            seq_fn='/data/plasmodium/pfalciparum/recon/data/raw/capillary/GATC/1181127/D07_PD0004-C-12891.fas',
            bad_qual_distance=None):
    read_group_info = "@RG\tID:%s_%s_%s\tSM:%s_%s_%s\tPL:capillary" % (ox_code, company, primers, ox_code, primers, company)
    bam_fn = "%s/%s_%s_%s.%s.bam" % (BAM_FILES_DIR, ox_code, company, primers, bad_qual_distance)
    
    fastq_fn = convert_to_fastq(seq_fn, bad_qual_distance=bad_qual_distance)
    
    bwa_command = "%s mem -M -R '%s' %s \"%s\" 2> /dev/null | samtools view -bSh - > %s 2> /dev/null" % (
        BWA_EXE, read_group_info, REF_GENOME, fastq_fn, bam_fn
    )
    !{bwa_command}
    
    samtools_command = "samtools sort %s %s" % (bam_fn, bam_fn.replace('.bam', ''))
    !{samtools_command}
    
    samtools_command = "samtools index %s" % bam_fn
    !{samtools_command}
    
    return(bam_fn)

# <codecell>

run_bwa()

# <codecell>

run_bwa(ox_code='PH0883-Cx')

# <codecell>

for bad_qual_distance in [None, 0, 1, 2]:
    gatc_fofn = GATC_BAM_FILES_FOFN % bad_qual_distance
    fo = open(gatc_fofn, 'w')
    for rec in tbl_gatc_metadata.selectne('ox_code', 'WATER').data():
        bam_fn = run_bwa(rec[1], 'GATC', rec[2], rec[0], bad_qual_distance=bad_qual_distance)
        print(bam_fn, file=fo)
    fo.close()

# <codecell>

# temp fix for PH0883-Cx
for bad_qual_distance in [None, 0, 1, 2]:
#     gatc_fofn = GATC_BAM_FILES_FOFN % bad_qual_distance
#     fo = open(gatc_fofn, 'w')
    for rec in tbl_gatc_metadata.selectne('ox_code', 'WATER').data():
        if rec[1] == 'PH0883-Cx':
            bam_fn = run_bwa(rec[1], 'GATC', rec[2], rec[0], bad_qual_distance=bad_qual_distance)
#             print(bam_fn, file=fo)
#     fo.close()

# <codecell>

def call_variants(bam_fn=GATC_BAM_FILES_FOFN % None, vcf_fn=GATC_CAPILLARY_VCF_FN % None):
    gatk_command = "%s \
 -T UnifiedGenotyper \
 -R %s \
 -I %s \
 -L Pf3D7_13_v3:1724943-1725954 \
 -ploidy 1 \
 -stand_call_conf 0 \
 -stand_emit_conf 0 \
 --min_base_quality_score 40 \
 -o %s" % (GATK_EXE, REF_GENOME, bam_fn, vcf_fn)
#  -o %s > /dev/null 2>&1" % (GATK_EXE, REF_GENOME, bam_fn, vcf_fn)
    print(gatk_command)
    !{gatk_command}

#  --genotype_likelihoods_model BOTH \
#  -minIndelCnt 1 \
#  --defaultBaseQualities 60 \

# <codecell>

calling_runs = {
    'GATC':[GATC_BAM_FILES_FOFN, GATC_CAPILLARY_VCF_FN],
}
for calling_run in calling_runs:
    for bad_qual_distance in [None, 0, 1, 2]:
        print(calling_run, bad_qual_distance)
        call_variants(calling_runs[calling_run][0] % bad_qual_distance, calling_runs[calling_run][1] % bad_qual_distance)

# <headingcell level=1>

# Create table of capillary calls

# <codecell>

wb = openpyxl.Workbook(optimized_write=True, encoding=locale.getpreferredencoding())

tbl_capillary_calls = collections.OrderedDict()
for callset in ['GATC']:
    tbl_capillary_calls[callset] = collections.OrderedDict()
    for bad_qual_distance in [2, 1, 0, None]:
        tbl_capillary_calls[callset][str(bad_qual_distance)] = (etl
            .fromvcf(calling_runs[callset][1] % bad_qual_distance)
            .vcfmeltsamples()
            .vcfunpackcall()
            .cut(['CHROM', 'POS', 'REF', 'ALT', 'SAMPLE', 'GT'])
            .addfield('variant', lambda rec: "%s__%s__%s__%s" % (rec['CHROM'], rec['POS'], rec['REF'], rec['ALT']))
            .pivot('variant', 'SAMPLE', 'GT', max)
        )
        print(callset, bad_qual_distance)
        ws = wb.create_sheet(title="%s_%s" % (callset, str(bad_qual_distance)))
        for row in tbl_capillary_calls[callset][str(bad_qual_distance)]:
            ws.append(row)
        tbl_capillary_calls[callset][str(bad_qual_distance)].displayall()

wb.save(CAPILLARY_CALLS_EXCEL_FN)

# <codecell>

for callset in ['GATC']:
    for bad_qual_distance in [None, 0, 1, 2]:
        print(callset, bad_qual_distance, len(tbl_capillary_calls[callset][str(bad_qual_distance)]))

# <headingcell level=1>

# Create table of merged calls from capillary, illumina and sequenom

# <codecell>

SQNM_VCF_FN

# <codecell>

tbl_capillary_samples

# <codecell>

wb = openpyxl.Workbook(optimized_write=True, encoding=locale.getpreferredencoding())

tbl_sequenom_calls = collections.OrderedDict()
callsets = collections.OrderedDict()
callsets['GATC'] = ['GATC']
# callsets['SB'] = ['Source_Bioscience']
# callsets['all'] = ['GATC', 'Source_Bioscience']
# callsets = {'GATC':['GATC'], 'SB':['Source_Bioscience'], 'all':['GATC', 'Source_Bioscience']}

for callset in callsets:
    tbl_sequenom_calls[callset] = (etl
        .fromvcf(SQNM_VCF_FN)
#         .select(lambda rec: rec['CHROM']=='Pf3D7_13_v3' and rec['POS'] >= 1724843 and rec['POS'] <= 1725958)
        .select(lambda rec: rec['CHROM']=='Pf3D7_13_v3' and rec['POS'] >= 1724943 and rec['POS'] <= 1725954)
        .vcfmeltsamples()
        .vcfunpackcall()
        .cut(['CHROM', 'POS', 'REF', 'ALT', 'SAMPLE', 'GT'])
        .addfield('variant', lambda rec: "%s__%s__%s__%s" % (rec['CHROM'], rec['POS'], rec['REF'], rec['ALT']))
        .selectin('SAMPLE', tbl_capillary_samples.values('ox_code'))
        .convert('SAMPLE', lambda val: "%s__SQNM" % val)
        .pivot('variant', 'SAMPLE', 'GT', max)
    )
    print(callset)
    ws = wb.create_sheet(title="%s" % (callset))
    for row in tbl_sequenom_calls[callset]:
        ws.append(row)
    tbl_sequenom_calls[callset].displayall()

wb.save(SEQUENOM_CALLS_EXCEL_FN)

# <codecell>

PGV4_VCF_FN

# <codecell>

vcf_reader = vcf.Reader(filename=PGV4_VCF_FN)
pgv4_samples_array = np.array(vcf_reader.samples)
# capillary_samples = tbl_capillary_samples.values('ox_code').array()
gatc_capillary_samples = tbl_capillary_samples.values('ox_code').array()
# sb_capillary_samples = tbl_capillary_samples.selecteq('company', 'Source_Bioscience').values('ox_code').array()

# common_samples = np.array(np.intersect1d(capillary_samples, pgv4_samples_array), dtype=[('sample', '<U45')])
gatc_common_samples = np.array(np.intersect1d(gatc_capillary_samples, pgv4_samples_array), dtype=[('sample', '<U45')])
# sb_common_samples = np.array(np.intersect1d(sb_capillary_samples, pgv4_samples_array), dtype=[('sample', '<U45')])
# etl.fromarray(common_samples).data().totsv(SAMPLES_FN % 'all')
etl.fromarray(gatc_common_samples).data().totsv(SAMPLES_FN % 'GATC')
# etl.fromarray(sb_common_samples).data().totsv(SAMPLES_FN % 'SB')

# <codecell>

len(tbl_capillary_samples.data())

# <codecell>

tbl_capillary_samples.valuecounts('is_in_pgv4_vcf')

# <codecell>

tbl_capillary_samples.selectnone('is_in_pgv4_vcf')

# <codecell>

SAMPLES_FN

# <codecell>

for callset in ['GATC']:
    output_fn = "%s/pgv4_gatk_subset_to_kelch13_less_DNA_%s.vcf.gz" % (DATA_DIR, callset)
    samples_fn = SAMPLES_FN % callset
    !{GATK_EXE} -T SelectVariants \
        -R {REF_GENOME} \
        -V {PGV4_VCF_FN} \
        -L Pf3D7_13_v3:1724943-1725954 \
        -o {output_fn} \
        --sample_file {samples_fn} \
        --unsafe LENIENT_VCF_PROCESSING
        
#         -L Pf3D7_13_v3:1724843-1725958 \

# <codecell>

etl.fromvcf("%s/pgv4_gatk_subset_to_kelch13_less_DNA_GATC.vcf.gz" % (DATA_DIR))

# <codecell>

tbl_illumina_calls = (etl
    .fromvcf("%s/pgv4_gatk_subset_to_kelch13_less_DNA_GATC.vcf.gz" % (DATA_DIR))
    .vcfmeltsamples()
    .vcfunpackcall()
    .cut(['CHROM', 'POS', 'REF', 'ALT', 'SAMPLE', 'AD'])
    .addfield('variant', lambda rec: "%s__%s__%s__%s" % (rec['CHROM'], rec['POS'], rec['REF'], rec['ALT']))
    .pivot('variant', 'SAMPLE', 'AD', max)
)
tbl_illumina_calls.displayall()

# <codecell>

def determine_GT(AD):
    if sum(AD) < 5:
        return(None)
    if len(AD) == 3:
        if AD[0] > 1 and AD[1] > 1:
            return("0/1")
        if AD[0] > 1 and AD[2] > 1:
            return("0/2")
        if AD[1] > 1 and AD[2] > 1:
            return("1/2")
    if(AD[0] > 1 and AD[1] > 1):
        return("0/1")
    if(AD[0] <= 1 and AD[1] > 1):
        return("1/1")
    if(AD[0] > 1 and AD[1] <= 1):
        return("0/0")
    else:
        return("unknown")
    

# <codecell>

def is_variable(rec):
    if all([val in ['0/0', None] for val in np.array(rec)[1:]]):
        return False
    else:
        return True

# <codecell>

wb = openpyxl.Workbook(optimized_write=True, encoding=locale.getpreferredencoding())

tbl_illumina_calls = collections.OrderedDict()
tbl_illumina_ADs = collections.OrderedDict()
for callset in ['GATC']:
    tbl_illumina_calls[callset] = (etl
        .fromvcf("%s/pgv4_gatk_subset_to_kelch13_less_DNA_%s.vcf.gz" % (DATA_DIR, callset))
        .vcfmeltsamples()
        .vcfunpackcall()
        .cut(['CHROM', 'POS', 'REF', 'ALT', 'SAMPLE', 'AD'])
        .addfield('variant', lambda rec: "%s__%s__%s__%s" % (rec['CHROM'], rec['POS'], rec['REF'], rec['ALT']))
        .addfield('GT', lambda rec: determine_GT(rec['AD']))
        .convert('SAMPLE', lambda val: "%s__ILMN" % val)
        .selectnotnone('GT')
        .pivot('variant', 'SAMPLE', 'GT', max)
        .select(is_variable)
    )
    ws = wb.create_sheet(title="GT_%s" % (callset))
    for row in tbl_illumina_calls[callset]:
        ws.append(row)
    tbl_illumina_calls[callset].displayall()

    tbl_illumina_ADs[callset] = (etl
        .fromvcf("%s/pgv4_gatk_subset_to_kelch13_less_DNA_%s.vcf.gz" % (DATA_DIR, callset))
        .vcfmeltsamples()
        .vcfunpackcall()
        .cut(['CHROM', 'POS', 'REF', 'ALT', 'SAMPLE', 'AD'])
        .addfield('variant', lambda rec: "%s__%s__%s__%s" % (rec['CHROM'], rec['POS'], rec['REF'], rec['ALT']))
        .convert('SAMPLE', lambda val: "%s__ILMN" % val)
        .convert('AD', str)
        .pivot('variant', 'SAMPLE', 'AD', max)
        .selectin('variant', tbl_illumina_calls[callset].values('variant'))
    )
    ws = wb.create_sheet(title="AD_%s" % (callset))
    for row in tbl_illumina_ADs[callset]:
        ws.append(row)
    tbl_illumina_ADs[callset].displayall()

wb.save(ILLUMINA_CALLS_EXCEL_FN)

# <codecell>

# Find which SNPs cannot be genotyped in which samples
(etl
    .fromvcf("%s/pgv4_gatk_subset_to_kelch13_less_DNA_GATC.vcf.gz" % (DATA_DIR))
    .vcfmeltsamples()
    .vcfunpackcall()
    .cut(['CHROM', 'POS', 'REF', 'ALT', 'SAMPLE', 'AD'])
    .addfield('variant', lambda rec: "%s__%s__%s__%s" % (rec['CHROM'], rec['POS'], rec['REF'], rec['ALT']))
    .addfield('GT', lambda rec: determine_GT(rec['AD']))
    .convert('SAMPLE', lambda val: "%s__ILMN" % val)
    .selectnone('GT')
).displayall()

# <codecell>

wb = openpyxl.Workbook(optimized_write=True, encoding=locale.getpreferredencoding())

tbl_merged_calls = collections.OrderedDict()
for callset in ['GATC']:
    tbl_merged_calls[callset] = collections.OrderedDict()
    for bad_qual_distance in [2, 1, 0, None]:
        tbl_merged_calls[callset][str(bad_qual_distance)] = (
            tbl_capillary_calls[callset][str(bad_qual_distance)]
            .outerjoin(tbl_sequenom_calls[callset])
            .outerjoin(tbl_illumina_calls[callset])
            .rename('variant', 'Concat_variant')
        )
        tbl_merged_calls[callset][str(bad_qual_distance)] = tbl_merged_calls[callset][str(bad_qual_distance)].cut(sort(tbl_merged_calls[callset][str(bad_qual_distance)].header()).tolist())
        ws = wb.create_sheet(title="%s_%s" % (callset, str(bad_qual_distance)))
        for row in tbl_merged_calls[callset][str(bad_qual_distance)]:
            ws.append(row)
        print(callset, str(bad_qual_distance))
        tbl_merged_calls[callset][str(bad_qual_distance)].displayall()

wb.save(MERGED_CALLS_EXCEL_FN)

# <codecell>

MERGED_CALLS_EXCEL_FN

# <codecell>


# <codecell>

# finding position of het call in PA0155-C (00AJ08.fas)
# blasted against 3D7
# start at pos 8 in read which maps to 1724996 in NCBI ref
# pos 1724996 in NCBI is 1724965 in our ref 

1724996 - 31

# <codecell>

# sanity check missed call in PH0227-C at 1725259
# should be at 1725290 with NCBI ref
1725259 + 31

# <codecell>


