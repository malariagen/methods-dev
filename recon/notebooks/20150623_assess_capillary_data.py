# -*- coding: utf-8 -*-
# <nbformat>3.0</nbformat>

# <headingcell level=1>

# Setup

# <codecell>

%run _shared_setup.ipynb

import openpyxl
import locale
from Bio import SeqIO
import petlx.bio

# <codecell>

GATC_EXCEL_FN = '/data/plasmodium/pfalciparum/recon/data/raw/capillary/GATC/GATC-Biotech seq barcode sample list.xlsx'
GATC_SEQ_DIRs = [
    '/data/plasmodium/pfalciparum/recon/data/raw/capillary/GATC/LIGHTrun 150527 raw data',
    '/data/plasmodium/pfalciparum/recon/data/raw/capillary/GATC/LIGHTrun 150605 raw data'
]

SB_EXCEL_FN = '/data/plasmodium/pfalciparum/recon/data/raw/capillary/Source_Bioscience/source_bioscience_sample_list_20150513.xls'
SB_SEQ_DIR = '/data/plasmodium/pfalciparum/recon/data/raw/capillary/Source_Bioscience/run_2015_05_13'

CAPILLARY_SAMPLES_FN = '/data/plasmodium/pfalciparum/recon/data/processed/capillary/initial_capillary_samples.xlsx'
!mkdir -p /data/plasmodium/pfalciparum/recon/data/processed/capillary

BAM_FILES_DIR = '/data/plasmodium/pfalciparum/recon/data/processed/capillary/bams'
ALL_BAM_FILES_FOFN = '/data/plasmodium/pfalciparum/recon/data/processed/capillary/bams/all_bams.%s.list'
GATC_BAM_FILES_FOFN = '/data/plasmodium/pfalciparum/recon/data/processed/capillary/bams/gatc_bams.%s.list'
SB_BAM_FILES_FOFN = '/data/plasmodium/pfalciparum/recon/data/processed/capillary/bams/sb_bams.%s.list'
!mkdir -p {BAM_FILES_DIR}

VARIANT_FILES_DIR = '/data/plasmodium/pfalciparum/recon/data/processed/capillary/variants'
!mkdir -p {VARIANT_FILES_DIR}
ALL_CAPILLARY_VCF_FN = "%s/all_capillary.%%s.vcf" % VARIANT_FILES_DIR
GATC_CAPILLARY_VCF_FN = "%s/gatc_capillary.%%s.vcf" % VARIANT_FILES_DIR
SB_CAPILLARY_VCF_FN = "%s/sb_capillary.%%s.vcf" % VARIANT_FILES_DIR
ILLUMINA_CALLS_EXCEL_FN = "%s/illumina_calls.xlsx" % VARIANT_FILES_DIR
SEQUENOM_CALLS_EXCEL_FN = "%s/sequenom_calls.xlsx" % VARIANT_FILES_DIR
CAPILLARY_CALLS_EXCEL_FN = "%s/capillary_calls.xlsx" % VARIANT_FILES_DIR
MERGED_CALLS_EXCEL_FN = "%s/merged_calls.xlsx" % VARIANT_FILES_DIR

SQNM_VCF_FN = DATA_DIR + '/Sqnm_data_DK1066_W1378_20150603.vcf.gz'
SAMPLES_FN = DATA_DIR + '/%s_capillary_samples.txt'

# <headingcell level=1>

# Parse sample details

# <codecell>

def determine_gatc_seq_fn(rec):
    if rec['primer_name'].startswith('DGGE'):
        return("%s/%s.fas" % (GATC_SEQ_DIRs[1], rec['seqID']))
    else:
        return("%s/%s.fas" % (GATC_SEQ_DIRs[0], rec['seqID']))
        

# <codecell>

def which_primers(rec):
    if rec['primer_name'].startswith('Forward'):
        return('Forward')
    if rec['primer_name'].startswith('Reverse'):
        return('Reverse')
    if rec['primer_name'].endswith('seq1'):
        return('seq1')
    if rec['primer_name'].endswith('seq2'):
        return('seq2')
    else:
        return('unknown')
    

# <codecell>

def convert_to_fastq(seq_fn='/data/plasmodium/pfalciparum/recon/data/raw/capillary/GATC/LIGHTrun 150527 raw data/00AJ08.fas',
                     fastq_fn=None, bad_qual_distance=None, qual_good=60, qual_bad=0):
    if fastq_fn is None:
        fastq_fn = seq_fn.replace('.fas', '.%s.fastq' % bad_qual_distance).replace('.seq', '.%s.fastq' % bad_qual_distance)
        
    # make fastq
    with open(seq_fn, "r") as fasta, open(fastq_fn, "w") as fastq:
        for record in SeqIO.parse(fasta, "fasta"):
            record.seq = record.seq.tomutable()
            record.letter_annotations["phred_quality"] = [qual_good] * len(record)
            if bad_qual_distance is not None:
                if bad_qual_distance==0:
                    for i in range(len(record)):
                        if record.seq[i].islower() or record.seq[i].upper()=='N':
                            record.letter_annotations["phred_quality"][i] = qual_bad
                elif bad_qual_distance==1:
                    for i in range(len(record)):
                        if i == 0:
                            if record.seq[i].islower() or record.seq[i+1].islower() or record.seq[i].upper()=='N':
                                record.letter_annotations["phred_quality"][i] = qual_bad
                        elif i == (len(record) - 1):
                            if record.seq[i].islower() or record.seq[i-1].islower() or record.seq[i].upper()=='N':
                                record.letter_annotations["phred_quality"][i] = qual_bad
                        else:
                            if record.seq[i].islower() or record.seq[i-1].islower() or record.seq[i+1].islower() or record.seq[i].upper()=='N':
                                record.letter_annotations["phred_quality"][i] = qual_bad
                        # Change N's to a, as bwa seems to ignore these and treat as deletion. These should have qual=0 and hence not get called as SNPs
                        if record.seq[i].upper()=='N':
                            record.seq[i] = 'a'
                elif bad_qual_distance==2:
                    for i in range(len(record)):
                        if i == 0:
                            if record.seq[i].islower() or record.seq[i+1].islower() or record.seq[i+2].islower() or record.seq[i].upper()=='N':
                                record.letter_annotations["phred_quality"][i] = qual_bad
                        if i == 1:
                            if record.seq[i].islower() or record.seq[i+1].islower() or record.seq[i-1].islower() or record.seq[i].upper()=='N':
                                record.letter_annotations["phred_quality"][i] = qual_bad
                        elif i == (len(record) - 1):
                            if record.seq[i].islower() or record.seq[i-1].islower() or record.seq[i-2].islower() or record.seq[i].upper()=='N':
                                record.letter_annotations["phred_quality"][i] = qual_bad
                        elif i == (len(record) - 2):
                            if record.seq[i].islower() or record.seq[i-1].islower() or record.seq[i+1].islower() or record.seq[i].upper()=='N':
                                record.letter_annotations["phred_quality"][i] = qual_bad
                        else:
                            if record.seq[i].islower() or record.seq[i-1].islower() or record.seq[i+1].islower() or record.seq[i-2].islower() or record.seq[i+2].islower() or record.seq[i].upper()=='N':
                                record.letter_annotations["phred_quality"][i] = qual_bad
                        # Change N's to a, as bwa seems to ignore these and treat as deletion. These should have qual=0 and hence not get called as SNPs
                        if record.seq[i].upper()=='N':
                            record.seq[i] = 'a'
                else:
                    print("Warning - can't handle bad_qual_distance=%s, assuming bad_qual_distance=None" % bad_qual_distance)
                        
            SeqIO.write(record, fastq, "fastq")
    
    return(fastq_fn)

convert_to_fastq()

# <codecell>

convert_to_fastq(bad_qual_distance=0)

# <codecell>

convert_to_fastq(bad_qual_distance=1)

# <codecell>

convert_to_fastq(bad_qual_distance=2)

# <codecell>

def sequence_length(seq_fn='/data/plasmodium/pfalciparum/recon/data/raw/capillary/GATC/LIGHTrun 150527 raw data/00AJ08.fas'):
    return(len(SeqIO.read(open(seq_fn, "r"), "fasta").seq))
sequence_length()

# <codecell>

tbl_gatc_metadata = (
    etl.fromxlsx(GATC_EXCEL_FN)
    .addfield('ox_code', lambda rec: rec['details'][0:8])
    .addfield('seq_fn', determine_gatc_seq_fn)
    .addfield('company', 'GATC')
    .addfield('primers', which_primers)
    .addfield('seq_length', lambda rec: sequence_length(rec['seq_fn']))
)
print(len(tbl_gatc_metadata))
tbl_gatc_metadata.displayall(index_header=True)

# <codecell>

tbl_sb_metadata = (
    etl.fromtsv(SB_EXCEL_FN)
    .selecteq('type', 'text')
    .rename('sample', 'ox_code')
    .addfield('seq_fn', lambda rec: "%s/%s" % (SB_SEQ_DIR, rec['file']))
    .addfield('company', 'Source_Bioscience')
    .addfield('seq_length', lambda rec: sequence_length(rec['seq_fn']))
)
print(len(tbl_sb_metadata))
tbl_sb_metadata.displayall(index_header=True)

# <codecell>

vcf_reader = vcf.Reader(filename=PGV4_VCF_FN)
pgv4_samples = etl.wrap([[x] for x in vcf_reader.samples]).pushheader(['ox_code']).addfield('is_in_pgv4_vcf', True)

# <codecell>

tbl_sample_sets

# <codecell>

tbl_pgv4_sample_manifest

# <codecell>

pgv4_samples

# <codecell>

tbl_capillary_samples = (
    tbl_gatc_metadata
    .cut(['ox_code', 'company'])
    .distinct('ox_code')
    .cat(tbl_sb_metadata.cut(['ox_code', 'company']).distinct('ox_code'))
    .leftjoin(tbl_sample_sets, lkey='ox_code', rkey='sample_code')
    .addfield('have_sequenom_data', lambda rec: rec['source_code'] is not None)
    .cut(['ox_code', 'company', 'have_sequenom_data'])
    .leftjoin(pgv4_samples)
    .leftjoin(tbl_pgv4_sample_manifest.selecteq('Exclude', 0), lkey='ox_code', rkey='Sample')
    .leftjoin(tbl_pgv4_sample_manifest, lkey='ox_code', rkey='Sample')
#     .cut(['ox_code', 'company', 'have_sequenom_data', 'is_in_pgv4_vcf', 'Pf3k v3'])
    .cut(['ox_code', 'company', 'have_sequenom_data', 'is_in_pgv4_vcf'])
    .replace('have_sequenom_data', False, None)
#     .convert('Pf3k v3', bool)
#     .replace('Pf3k v3', False, None)
    .leftjoin(tbl_gatc_metadata.selecteq('primer_name', 'Forward-DK_sqnm_12891').cut(['ox_code', 'seq_length']), lkey='ox_code', rkey='ox_code')
    .rename('seq_length', 'length_gatc')
    .leftjoin(tbl_sb_metadata.selecteq('primer_name', 'Forward - DK_sqnm_12891').cut(['ox_code', 'seq_length']), lkey='ox_code', rkey='ox_code')
    .rename('seq_length', 'length_sb')
    .addfield('forward', lambda rec: rec['length_gatc'] if rec['length_gatc'] is not None else rec['length_sb'])
    .cutout('length_gatc', 'length_sb')
    .leftjoin(tbl_gatc_metadata.selecteq('primer_name', 'Reverse-DK_sqnm_12825').cut(['ox_code', 'seq_length']), lkey='ox_code', rkey='ox_code')
    .rename('seq_length', 'length_gatc')
    .leftjoin(tbl_sb_metadata.selecteq('primer_name', 'Reverse - DK_sqnm_12825').cut(['ox_code', 'seq_length']), lkey='ox_code', rkey='ox_code')
    .rename('seq_length', 'length_sb')
    .addfield('reverse', lambda rec: rec['length_gatc'] if rec['length_gatc'] is not None else rec['length_sb'])
    .cutout('length_gatc', 'length_sb')
    .leftjoin(tbl_gatc_metadata.selecteq('primer_name', 'DGGE_12825_seq1').cut(['ox_code', 'seq_length']), lkey='ox_code', rkey='ox_code')
    .rename('seq_length', 'DGGE_seq1')
    .leftjoin(tbl_gatc_metadata.selecteq('primer_name', 'DGGE_12825_seq2').cut(['ox_code', 'seq_length']), lkey='ox_code', rkey='ox_code')
    .rename('seq_length', 'DGGE_seq2')
    .sort(['company', 'ox_code'])
)
print(len(tbl_capillary_samples.data()))
tbl_capillary_samples.displayall()

# <codecell>

tbl_capillary_samples.toxlsx(CAPILLARY_SAMPLES_FN)

# <codecell>

!{BWA_EXE} index {REF_GENOME}

# <codecell>

def run_bwa(ox_code='PA0155-C', company='GATC', primers='Forward',
            seq_fn='/data/plasmodium/pfalciparum/recon/data/raw/capillary/GATC/LIGHTrun 150527 raw data/00AJ08.fas',
            bad_qual_distance=None):
    read_group_info = "@RG\tID:%s_%s_%s\tSM:%s_%s_%s\tPL:capillary" % (ox_code, company, primers, ox_code, primers, company)
    bam_fn = "%s/%s_%s_%s.%s.bam" % (BAM_FILES_DIR, ox_code, company, primers, bad_qual_distance)
    
    fastq_fn = convert_to_fastq(seq_fn, bad_qual_distance=bad_qual_distance)
    
    bwa_command = "%s mem -M -R '%s' %s \"%s\" 2> /dev/null | samtools view -bSh - > %s 2> /dev/null" % (
        BWA_EXE, read_group_info, REF_GENOME, fastq_fn, bam_fn
    )
    !{bwa_command}
    
    samtools_command = "samtools sort %s %s" % (bam_fn, bam_fn.replace('.bam', ''))
    !{samtools_command}
    
    samtools_command = "samtools index %s" % bam_fn
    !{samtools_command}
    
    return(bam_fn)

# <codecell>

run_bwa()

# <codecell>

for bad_qual_distance in [None, 0, 1, 2]:
    gatc_fofn = GATC_BAM_FILES_FOFN % bad_qual_distance
    fo = open(gatc_fofn, 'w')
    for rec in tbl_gatc_metadata.data():
        bam_fn = run_bwa(rec[8], rec[10], rec[11], rec[9], bad_qual_distance=bad_qual_distance)
        print(bam_fn, file=fo)
    fo.close()

    sb_fofn = SB_BAM_FILES_FOFN % bad_qual_distance
    fo = open(sb_fofn, 'w')
    for rec in tbl_sb_metadata.data():
        bam_fn = run_bwa(rec[3], rec[8], rec[4], rec[7], bad_qual_distance=bad_qual_distance)
        print(bam_fn, file=fo)
    fo.close()
    
    all_fofn = ALL_BAM_FILES_FOFN % bad_qual_distance
    !cat {gatc_fofn} {sb_fofn} > {all_fofn}

# <codecell>

def call_variants(bam_fn=ALL_BAM_FILES_FOFN % None, vcf_fn=ALL_CAPILLARY_VCF_FN % None):
    gatk_command = "%s \
 -T UnifiedGenotyper \
 -R %s \
 -I %s \
 -L Pf3D7_13_v3:1724943-1725954 \
 -ploidy 1 \
 -stand_call_conf 0 \
 -stand_emit_conf 0 \
 --min_base_quality_score 40 \
 -o %s > /dev/null 2>&1" % (GATK_EXE, REF_GENOME, bam_fn, vcf_fn)
    print(gatk_command)
    !{gatk_command}

#  --genotype_likelihoods_model BOTH \
#  -minIndelCnt 1 \
#  --defaultBaseQualities 60 \

# <codecell>

calling_runs = {
    'all':[ALL_BAM_FILES_FOFN, ALL_CAPILLARY_VCF_FN],
    'GATC':[GATC_BAM_FILES_FOFN, GATC_CAPILLARY_VCF_FN],
    'SB':[SB_BAM_FILES_FOFN, SB_CAPILLARY_VCF_FN],
}
for calling_run in calling_runs:
    for bad_qual_distance in [None, 0, 1, 2]:
        print(calling_run, bad_qual_distance)
        call_variants(calling_runs[calling_run][0] % bad_qual_distance, calling_runs[calling_run][1] % bad_qual_distance)

# <headingcell level=1>

# Create table of capillary calls

# <codecell>

wb = openpyxl.Workbook(optimized_write=True, encoding=locale.getpreferredencoding())

tbl_capillary_calls = collections.OrderedDict()
for callset in ['GATC', 'SB', 'all']:
    tbl_capillary_calls[callset] = collections.OrderedDict()
    for bad_qual_distance in [2, 1, 0, None]:
        tbl_capillary_calls[callset][str(bad_qual_distance)] = (etl
            .fromvcf(calling_runs[callset][1] % bad_qual_distance)
            .vcfmeltsamples()
            .vcfunpackcall()
            .cut(['CHROM', 'POS', 'REF', 'ALT', 'SAMPLE', 'GT'])
            .addfield('variant', lambda rec: "%s__%s__%s__%s" % (rec['CHROM'], rec['POS'], rec['REF'], rec['ALT']))
            .pivot('variant', 'SAMPLE', 'GT', max)
        )
        print(callset, bad_qual_distance)
        ws = wb.create_sheet(title="%s_%s" % (callset, str(bad_qual_distance)))
        for row in tbl_capillary_calls[callset][str(bad_qual_distance)]:
            ws.append(row)
        tbl_capillary_calls[callset][str(bad_qual_distance)].displayall()

wb.save(CAPILLARY_CALLS_EXCEL_FN)

# <codecell>

for callset in ['all', 'GATC', 'SB']:
    for bad_qual_distance in [None, 0, 1, 2]:
        print(callset, bad_qual_distance, len(tbl_capillary_calls[callset][str(bad_qual_distance)]))

# <headingcell level=1>

# Create table of merged calls from capillary, illumina and sequenom

# <codecell>

SQNM_VCF_FN

# <codecell>

wb = openpyxl.Workbook(optimized_write=True, encoding=locale.getpreferredencoding())

tbl_sequenom_calls = collections.OrderedDict()
callsets = collections.OrderedDict()
callsets['GATC'] = ['GATC']
callsets['SB'] = ['Source_Bioscience']
callsets['all'] = ['GATC', 'Source_Bioscience']
# callsets = {'GATC':['GATC'], 'SB':['Source_Bioscience'], 'all':['GATC', 'Source_Bioscience']}

for callset in callsets:
    tbl_sequenom_calls[callset] = (etl
        .fromvcf(SQNM_VCF_FN)
#         .select(lambda rec: rec['CHROM']=='Pf3D7_13_v3' and rec['POS'] >= 1724843 and rec['POS'] <= 1725958)
        .select(lambda rec: rec['CHROM']=='Pf3D7_13_v3' and rec['POS'] >= 1724943 and rec['POS'] <= 1725954)
        .vcfmeltsamples()
        .vcfunpackcall()
        .cut(['CHROM', 'POS', 'REF', 'ALT', 'SAMPLE', 'GT'])
        .addfield('variant', lambda rec: "%s__%s__%s__%s" % (rec['CHROM'], rec['POS'], rec['REF'], rec['ALT']))
        .selectin('SAMPLE', tbl_capillary_samples.selectin('company', callsets[callset]).values('ox_code'))
        .convert('SAMPLE', lambda val: "%s__SQNM" % val)
        .pivot('variant', 'SAMPLE', 'GT', max)
    )
    print(callset)
    ws = wb.create_sheet(title="%s" % (callset))
    for row in tbl_sequenom_calls[callset]:
        ws.append(row)
    tbl_sequenom_calls[callset].displayall()

wb.save(SEQUENOM_CALLS_EXCEL_FN)

# <codecell>

PGV4_VCF_FN

# <codecell>

vcf_reader = vcf.Reader(filename=PGV4_VCF_FN)
pgv4_samples = np.array(vcf_reader.samples)
capillary_samples = tbl_capillary_samples.values('ox_code').array()
gatc_capillary_samples = tbl_capillary_samples.selecteq('company', 'GATC').values('ox_code').array()
sb_capillary_samples = tbl_capillary_samples.selecteq('company', 'Source_Bioscience').values('ox_code').array()

common_samples = np.array(np.intersect1d(capillary_samples, pgv4_samples), dtype=[('sample', '<U45')])
gatc_common_samples = np.array(np.intersect1d(gatc_capillary_samples, pgv4_samples), dtype=[('sample', '<U45')])
sb_common_samples = np.array(np.intersect1d(sb_capillary_samples, pgv4_samples), dtype=[('sample', '<U45')])
etl.fromarray(common_samples).data().totsv(SAMPLES_FN % 'all')
etl.fromarray(gatc_common_samples).data().totsv(SAMPLES_FN % 'GATC')
etl.fromarray(sb_common_samples).data().totsv(SAMPLES_FN % 'SB')

# <codecell>

for callset in ['all', 'GATC', 'SB']:
    output_fn = "%s/pgv4_gatk_subset_to_kelch13_%s.vcf.gz" % (DATA_DIR, callset)
    samples_fn = SAMPLES_FN % callset
    !{GATK_EXE} -T SelectVariants \
        -R {REF_GENOME} \
        -V {PGV4_VCF_FN} \
        -L Pf3D7_13_v3:1724943-1725954 \
        -o {output_fn} \
        --sample_file {samples_fn} \
        --unsafe LENIENT_VCF_PROCESSING
        
#         -L Pf3D7_13_v3:1724843-1725958 \

# <codecell>

etl.fromvcf("%s/pgv4_gatk_subset_to_kelch13_GATC.vcf.gz" % (DATA_DIR))

# <codecell>

tbl_illumina_calls = (etl
    .fromvcf("%s/pgv4_gatk_subset_to_kelch13.vcf.gz" % (DATA_DIR))
    .vcfmeltsamples()
    .vcfunpackcall()
    .cut(['CHROM', 'POS', 'REF', 'ALT', 'SAMPLE', 'AD'])
    .addfield('variant', lambda rec: "%s__%s__%s__%s" % (rec['CHROM'], rec['POS'], rec['REF'], rec['ALT']))
    .pivot('variant', 'SAMPLE', 'AD', max)
)
tbl_illumina_calls.displayall()

# <codecell>

def determine_GT(AD):
    if sum(AD) < 5:
        return(None)
    if len(AD) == 3:
        if AD[0] > 1 and AD[1] > 1:
            return("0/1")
        if AD[0] > 1 and AD[2] > 1:
            return("0/2")
        if AD[1] > 1 and AD[2] > 1:
            return("1/2")
    if(AD[0] > 1 and AD[1] > 1):
        return("0/1")
    if(AD[0] <= 1 and AD[1] > 1):
        return("1/1")
    if(AD[0] > 1 and AD[1] <= 1):
        return("0/0")
    else:
        return("unknown")
    

# <codecell>

def is_variable(rec):
    if all([val in ['0/0', None] for val in np.array(rec)[1:]]):
        return False
    else:
        return True

# <codecell>

wb = openpyxl.Workbook(optimized_write=True, encoding=locale.getpreferredencoding())

tbl_illumina_calls = collections.OrderedDict()
tbl_illumina_ADs = collections.OrderedDict()
for callset in ['GATC', 'SB', 'all']:
    tbl_illumina_calls[callset] = (etl
        .fromvcf("%s/pgv4_gatk_subset_to_kelch13_%s.vcf.gz" % (DATA_DIR, callset))
        .vcfmeltsamples()
        .vcfunpackcall()
        .cut(['CHROM', 'POS', 'REF', 'ALT', 'SAMPLE', 'AD'])
        .addfield('variant', lambda rec: "%s__%s__%s__%s" % (rec['CHROM'], rec['POS'], rec['REF'], rec['ALT']))
        .addfield('GT', lambda rec: determine_GT(rec['AD']))
        .convert('SAMPLE', lambda val: "%s__ILMN" % val)
        .selectnotnone('GT')
        .pivot('variant', 'SAMPLE', 'GT', max)
        .select(is_variable)
    )
    ws = wb.create_sheet(title="GT_%s" % (callset))
    for row in tbl_illumina_calls[callset]:
        ws.append(row)
    tbl_illumina_calls[callset].displayall()

    tbl_illumina_ADs[callset] = (etl
        .fromvcf("%s/pgv4_gatk_subset_to_kelch13_%s.vcf.gz" % (DATA_DIR, callset))
        .vcfmeltsamples()
        .vcfunpackcall()
        .cut(['CHROM', 'POS', 'REF', 'ALT', 'SAMPLE', 'AD'])
        .addfield('variant', lambda rec: "%s__%s__%s__%s" % (rec['CHROM'], rec['POS'], rec['REF'], rec['ALT']))
        .convert('SAMPLE', lambda val: "%s__ILMN" % val)
        .convert('AD', str)
        .pivot('variant', 'SAMPLE', 'AD', max)
        .selectin('variant', tbl_illumina_calls[callset].values('variant'))
    )
    ws = wb.create_sheet(title="AD_%s" % (callset))
    for row in tbl_illumina_ADs[callset]:
        ws.append(row)
    tbl_illumina_ADs[callset].displayall()

wb.save(ILLUMINA_CALLS_EXCEL_FN)

# <codecell>

# Find which SNPs cannot be genotyped in which samples
(etl
    .fromvcf("%s/pgv4_gatk_subset_to_kelch13.vcf.gz" % (DATA_DIR))
    .vcfmeltsamples()
    .vcfunpackcall()
    .cut(['CHROM', 'POS', 'REF', 'ALT', 'SAMPLE', 'AD'])
    .addfield('variant', lambda rec: "%s__%s__%s__%s" % (rec['CHROM'], rec['POS'], rec['REF'], rec['ALT']))
    .addfield('GT', lambda rec: determine_GT(rec['AD']))
    .convert('SAMPLE', lambda val: "%s__ILMN" % val)
    .selectnone('GT')
).displayall()

# <codecell>

wb = openpyxl.Workbook(optimized_write=True, encoding=locale.getpreferredencoding())

tbl_merged_calls = collections.OrderedDict()
for callset in ['GATC', 'SB', 'all']:
    tbl_merged_calls[callset] = collections.OrderedDict()
    for bad_qual_distance in [2, 1, 0, None]:
        tbl_merged_calls[callset][str(bad_qual_distance)] = (
            tbl_capillary_calls[callset][str(bad_qual_distance)]
            .outerjoin(tbl_sequenom_calls[callset])
            .outerjoin(tbl_illumina_calls[callset])
            .rename('variant', 'Concat_variant')
        )
        tbl_merged_calls[callset][str(bad_qual_distance)] = tbl_merged_calls[callset][str(bad_qual_distance)].cut(sort(tbl_merged_calls[callset][str(bad_qual_distance)].header()).tolist())
        ws = wb.create_sheet(title="%s_%s" % (callset, str(bad_qual_distance)))
        for row in tbl_merged_calls[callset][str(bad_qual_distance)]:
            ws.append(row)
        print(callset, str(bad_qual_distance))
        tbl_merged_calls[callset][str(bad_qual_distance)].displayall()

wb.save(MERGED_CALLS_EXCEL_FN)

# <codecell>


# <codecell>

# finding position of het call in PA0155-C (00AJ08.fas)
# blasted against 3D7
# start at pos 8 in read which maps to 1724996 in NCBI ref
# pos 1724996 in NCBI is 1724965 in our ref 

1724996 - 31

# <codecell>

# sanity check missed call in PH0227-C at 1725259
# should be at 1725290 with NCBI ref
1725259 + 31

# <codecell>


